#!/usr/bin/env python3
"""
Generate fake DevCo KPI submission files for Whitehelmet end-to-end testing.

Produces (in Project Documents/KPI test sheets/Test Submissions/):
  - 4 x Safety KPI submissions   (filled KPI Template.xlsx)
  - 4 x Quality KPI submissions  (filled Quality template)
  - 1 x Master KPI Consolidation reference sheet (Safety + Quality tabs)

Run from repo root:
    cd /path/to/whitehelmet
    python generate_test_kpi.py
"""

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# ── Paths ──────────────────────────────────────────────────────────────────────
REPO_ROOT    = Path(__file__).parent
KPI_DIR      = REPO_ROOT / "Project Documents" / "KPI test sheets"
SAFETY_TPL   = KPI_DIR / "Safety"  / "KPI Template.xlsx"
QUALITY_TPL  = KPI_DIR / "Quality" / "DevCo Name - KPI - MONTH-YEAR-Template.xlsx"
OUTPUT_DIR   = KPI_DIR / "Test Submissions"

MONTH_YEAR       = "APR-2026"
DATA_DATE        = date(2026, 4, 30)
SUBMISSION_DATE  = date(2026, 4, 30)

# ── Test DevCo data (4 companies, varied performance levels) ───────────────────
#
# Safety input columns F-O (rows 10-14 in ARD sheet, one row per sub-project):
#   F=manhours, G=safe_manhours, H=fatalities, I=lti_incidents, J=lti_manhours,
#   K=recordable, L=near_miss, M=safety_obs, N=leadership_walks, O=recognitions
#
# Quality input columns B-Q (rows 10-12 in Quality sheet, one row per WBS item):
#   B=Level1, C=Level2, D=Level3,
#   E=recv_submittals, F=code_a, G=code_b, H=code_c,
#   I=planned_audits, J=conducted_audits, K=leadership_visits,
#   L=mirs_received, M=mirs_accepted, N=wirs_received, O=wirs_accepted,
#   P=ncrs_raised, Q=ncrs_closed
# ──────────────────────────────────────────────────────────────────────────────

DEVCOS = [
    {
        "short": "FMSCO",
        "name":  "First Middle East Company (FMSCO)",
        "profile": "Excellent",
        # Safety: one project row (row 10), data in columns E-O
        "safety_rows": [
            dict(project="THE LINE – Excavation M44-M46",
                 manhours=520_000, safe_manhours=519_900,
                 fatalities=0, lti_incidents=0, lti_manhours=0,
                 recordable=0, near_miss=310, safety_obs=680,
                 leadership_walks=48, recognitions=245),
        ],
        # Quality: 3 WBS sub-rows (rows 10-12)
        "quality_rows": [
            dict(lvl1="Earthworks", lvl2="Excavation", lvl3="Zone A",
                 recv=42, ca=36, cb=4, cc=2,
                 p_aud=8, c_aud=8, lv=2,
                 mirs_r=22, mirs_a=21, wirs_r=13, wirs_a=13, ncrs_r=4, ncrs_c=4),
            dict(lvl1="Earthworks", lvl2="Excavation", lvl3="Zone B",
                 recv=38, ca=32, cb=5, cc=1,
                 p_aud=7, c_aud=7, lv=3,
                 mirs_r=20, mirs_a=20, wirs_r=12, wirs_a=12, ncrs_r=3, ncrs_c=3),
            dict(lvl1="Earthworks", lvl2="Backfill",   lvl3="Zone A",
                 recv=35, ca=30, cb=3, cc=2,
                 p_aud=7, c_aud=6, lv=2,
                 mirs_r=16, mirs_a=15, wirs_r=9,  wirs_a=8,  ncrs_r=3, ncrs_c=3),
        ],
    },
    {
        "short": "AlMardoof",
        "name":  "Al Mardoof Construction Co.",
        "profile": "Good",
        "safety_rows": [
            dict(project="THE LINE – Excavation WO-01 & WO-02",
                 manhours=280_000, safe_manhours=279_850,
                 fatalities=0, lti_incidents=1, lti_manhours=150,
                 recordable=1, near_miss=120, safety_obs=290,
                 leadership_walks=22, recognitions=95),
        ],
        "quality_rows": [
            dict(lvl1="Earthworks", lvl2="Excavation", lvl3="Zone C",
                 recv=30, ca=25, cb=3, cc=2,
                 p_aud=6, c_aud=5, lv=2,
                 mirs_r=15, mirs_a=14, wirs_r=9, wirs_a=8, ncrs_r=3, ncrs_c=2),
            dict(lvl1="Earthworks", lvl2="Backfill",   lvl3="Zone B",
                 recv=28, ca=22, cb=4, cc=2,
                 p_aud=5, c_aud=4, lv=1,
                 mirs_r=14, mirs_a=12, wirs_r=8, wirs_a=7, ncrs_r=3, ncrs_c=2),
            dict(lvl1="Earthworks", lvl2="Grading",    lvl3="Zone C",
                 recv=27, ca=23, cb=3, cc=1,
                 p_aud=5, c_aud=5, lv=1,
                 mirs_r=13, mirs_a=12, wirs_r=8, wirs_a=7, ncrs_r=2, ncrs_c=2),
        ],
    },
    {
        "short": "Enjaz",
        "name":  "Enjaz Contracting Co.",
        "profile": "Average",
        "safety_rows": [
            dict(project="THE LINE – Excavation Zone 4",
                 manhours=175_000, safe_manhours=174_200,
                 fatalities=0, lti_incidents=2, lti_manhours=800,
                 recordable=3, near_miss=35, safety_obs=95,
                 leadership_walks=9, recognitions=22),
        ],
        "quality_rows": [
            dict(lvl1="Piling", lvl2="Bored Piles", lvl3="Section A",
                 recv=22, ca=16, cb=4, cc=2,
                 p_aud=5, c_aud=3, lv=1,
                 mirs_r=11, mirs_a=9,  wirs_r=6, wirs_a=5, ncrs_r=3, ncrs_c=2),
            dict(lvl1="Piling", lvl2="Bored Piles", lvl3="Section B",
                 recv=20, ca=14, cb=4, cc=2,
                 p_aud=4, c_aud=3, lv=1,
                 mirs_r=10, mirs_a=8,  wirs_r=6, wirs_a=4, ncrs_r=2, ncrs_c=1),
            dict(lvl1="Piling", lvl2="Pile Caps",   lvl3="Section A",
                 recv=18, ca=14, cb=3, cc=1,
                 p_aud=3, c_aud=3, lv=0,
                 mirs_r=9,  mirs_a=7,  wirs_r=6, wirs_a=5, ncrs_r=1, ncrs_c=1),
        ],
    },
    {
        "short": "AFAC",
        "name":  "AFAC Construction Group",
        "profile": "Good",
        "safety_rows": [
            dict(project="THE LINE – Excavation Zone 5",
                 manhours=320_000, safe_manhours=319_950,
                 fatalities=0, lti_incidents=0, lti_manhours=0,
                 recordable=1, near_miss=165, safety_obs=380,
                 leadership_walks=28, recognitions=112),
        ],
        "quality_rows": [
            dict(lvl1="Earthworks", lvl2="Excavation", lvl3="Zone E",
                 recv=32, ca=28, cb=3, cc=1,
                 p_aud=7, c_aud=6, lv=2,
                 mirs_r=17, mirs_a=17, wirs_r=10, wirs_a=10, ncrs_r=3, ncrs_c=3),
            dict(lvl1="Earthworks", lvl2="Backfill",   lvl3="Zone D",
                 recv=30, ca=27, cb=2, cc=1,
                 p_aud=6, c_aud=6, lv=2,
                 mirs_r=16, mirs_a=16, wirs_r=10, wirs_a=10, ncrs_r=2, ncrs_c=2),
            dict(lvl1="Earthworks", lvl2="Grading",    lvl3="Zone E",
                 recv=30, ca=25, cb=4, cc=1,
                 p_aud=5, c_aud=5, lv=1,
                 mirs_r=15, mirs_a=14, wirs_r=8,  wirs_a=7,  ncrs_r=2, ncrs_c=2),
        ],
    },
]


# ── File generators ────────────────────────────────────────────────────────────

def fill_safety(devco: dict, output_path: Path) -> None:
    """Copy safety template, fill company info and ARD data rows."""
    wb = openpyxl.load_workbook(SAFETY_TPL)

    # Company Information
    ci = wb["Company Infomation"]
    ci["D3"] = devco["name"]
    ci["D4"] = DATA_DATE
    ci["D5"] = SUBMISSION_DATE

    # ARD sheet — data rows 10-14 (template is fully empty in those rows)
    ard = wb["ARD"]
    for i, s in enumerate(devco["safety_rows"]):
        r = 10 + i
        ard.cell(r, 5).value  = s["project"]
        ard.cell(r, 6).value  = s["manhours"]
        ard.cell(r, 7).value  = s["safe_manhours"]
        ard.cell(r, 8).value  = s["fatalities"]
        ard.cell(r, 9).value  = s["lti_incidents"]
        ard.cell(r, 10).value = s["lti_manhours"]
        ard.cell(r, 11).value = s["recordable"]
        ard.cell(r, 12).value = s["near_miss"]
        ard.cell(r, 13).value = s["safety_obs"]
        ard.cell(r, 14).value = s["leadership_walks"]
        ard.cell(r, 15).value = s["recognitions"]
        # Row 16 SUM formulas + Row 17 rate formulas auto-calculate in Excel

    wb.save(output_path)


def fill_quality(devco: dict, output_path: Path) -> None:
    """Copy quality template, fill company info and Quality data rows."""
    wb = openpyxl.load_workbook(QUALITY_TPL)

    # Company Information
    ci = wb["Company Infomation"]
    ci["D3"] = devco["name"]
    ci["D4"] = DATA_DATE
    ci["D5"] = SUBMISSION_DATE

    # Quality sheet — data rows 10-12 (col A already has seq numbers 1,2,3)
    q = wb["Quality"]
    for i, d in enumerate(devco["quality_rows"]):
        r = 10 + i
        q.cell(r, 2).value  = d["lvl1"]
        q.cell(r, 3).value  = d["lvl2"]
        q.cell(r, 4).value  = d["lvl3"]
        q.cell(r, 5).value  = d["recv"]
        q.cell(r, 6).value  = d["ca"]
        q.cell(r, 7).value  = d["cb"]
        q.cell(r, 8).value  = d["cc"]
        q.cell(r, 9).value  = d["p_aud"]
        q.cell(r, 10).value = d["c_aud"]
        q.cell(r, 11).value = d["lv"]
        q.cell(r, 12).value = d["mirs_r"]
        q.cell(r, 13).value = d["mirs_a"]   # M: Accepted MIRs (input, not formula)
        q.cell(r, 14).value = d["wirs_r"]
        q.cell(r, 15).value = d["wirs_a"]   # O: Accepted WIRs (input, not formula)
        q.cell(r, 16).value = d["ncrs_r"]
        q.cell(r, 17).value = d["ncrs_c"]   # Q: Closed NCRs  (input, not formula)

    wb.save(output_path)


# ── Master KPI reference sheet ─────────────────────────────────────────────────

def _rate(count: int, manhours: int) -> float:
    return round(count * 200_000 / manhours, 4) if manhours else 0.0


def _safety_composite(fr, ltifr, trir, nmr, sor, ler, ipr) -> float:
    """Apply PIF scoring thresholds and weights from Formula.xlsx."""
    def s_fr(v):    return 100 if v == 0 else 90 if v <= 0.01 else 80 if v <= 0.05 else 70 if v <= 0.1 else 0
    def s_ltifr(v): return 100 if v == 0 else 90 if v <= 0.04 else 80 if v <= 0.10 else 70 if v <= 0.20 else 0
    def s_trir(v):  return 100 if v == 0 else 90 if v <= 0.045 else 80 if v <= 0.09 else 70 if v <= 0.15 else 0
    def s_nm(v):    return 100 if v >= 40 else 90 if v >= 30 else 80 if v >= 20 else 70 if v >= 10 else 60
    def s_so(v):    return 100 if v >= 100 else 90 if v >= 80 else 80 if v >= 60 else 70 if v >= 40 else 60
    def s_le(v):    return 100 if v >= 7.0 else 95 if v >= 5.0 else 90 if v >= 3.0 else 85 if v >= 2.0 else 80
    def s_ip(v):    return 100 if v >= 0.4 else 95 if v >= 0.3 else 90 if v >= 0.2 else 85 if v >= 0.1 else 80
    # Weights: FR=0.20, LTIFR=0.10, TRIR=0.10, NM=0.15, SO=0.15, LE=0.15, IP=0.15
    return round(
        s_fr(fr)    * 0.20 + s_ltifr(ltifr) * 0.10 + s_trir(trir) * 0.10
        + s_nm(nmr) * 0.15 + s_so(sor)      * 0.15 + s_le(ler)    * 0.15
        + s_ip(ipr) * 0.15, 2
    )


def make_master() -> None:
    """Generate master KPI consolidation reference xlsx (Safety + Quality tabs)."""
    wb = openpyxl.Workbook()

    # ── Tab 1: Safety KPIs ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Safety KPIs"
    ws.column_dimensions["A"].width = 36
    for col in "BCDEFGHIJKLMNOPQRS":
        ws.column_dimensions[col].width = 18

    safety_headers = [
        "DevCo", "Reporting Month",
        "Total Manhours", "Safe Manhours", "Fatalities",
        "LTI Incidents", "LTI Manhours", "Recordable Incidents",
        "Near Misses", "Safety Observations", "Leadership Walks", "Recognitions",
        "Fatality Rate (FR)", "LTIFR", "TRIR",
        "Near Miss Rate", "Safety Obs Rate", "Leadership Rate", "Incentive Rate",
        "Weighted Safety Score (%)",
    ]
    ws.append(safety_headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for d in DEVCOS:
        # Aggregate safety rows (supports multiple project sub-rows)
        mh   = sum(r["manhours"]         for r in d["safety_rows"])
        smh  = sum(r["safe_manhours"]    for r in d["safety_rows"])
        fat  = sum(r["fatalities"]       for r in d["safety_rows"])
        lti  = sum(r["lti_incidents"]    for r in d["safety_rows"])
        ltih = sum(r["lti_manhours"]     for r in d["safety_rows"])
        rec  = sum(r["recordable"]       for r in d["safety_rows"])
        nm   = sum(r["near_miss"]        for r in d["safety_rows"])
        so   = sum(r["safety_obs"]       for r in d["safety_rows"])
        lw   = sum(r["leadership_walks"] for r in d["safety_rows"])
        rew  = sum(r["recognitions"]     for r in d["safety_rows"])

        fr    = _rate(fat, mh)
        ltifr = _rate(lti, mh)
        trir  = _rate(rec, mh)
        nmr   = _rate(nm,  mh)
        sor   = _rate(so,  mh)
        ler   = _rate(lw,  mh)
        ipr   = _rate(rew, mh)
        score = _safety_composite(fr, ltifr, trir, nmr, sor, ler, ipr)

        ws.append([
            d["name"], MONTH_YEAR,
            mh, smh, fat, lti, ltih, rec, nm, so, lw, rew,
            fr, ltifr, trir, nmr, sor, ler, ipr, score,
        ])

    # ── Tab 2: Quality KPIs ────────────────────────────────────────────────────
    wq = wb.create_sheet("Quality KPIs")
    wq.column_dimensions["A"].width = 36
    for col in "BCDEFGHIJKLMNOPQRSTU":
        wq.column_dimensions[col].width = 18

    quality_headers = [
        "DevCo", "Reporting Month",
        "Total Submittals Received", "Code A", "Code B", "Code C",
        "Code A Approval Rate (%)",
        "Planned Audits", "Conducted Audits", "Audit Completion Rate (%)",
        "Leadership Visits",
        "MIRs Received", "MIRs Accepted", "MIR Acceptance Rate (%)",
        "WIRs Received", "WIRs Accepted", "WIR Acceptance Rate (%)",
        "NCRs Raised", "NCRs Closed", "NCR Closure Rate (%)",
        "Weighted Quality Score (%)",
    ]
    wq.append(quality_headers)
    for cell in wq[1]:
        cell.font = Font(bold=True)

    for d in DEVCOS:
        rows   = d["quality_rows"]
        recv   = sum(r["recv"]  for r in rows)
        ca     = sum(r["ca"]    for r in rows)
        cb     = sum(r["cb"]    for r in rows)
        cc     = sum(r["cc"]    for r in rows)
        p_aud  = sum(r["p_aud"] for r in rows)
        c_aud  = sum(r["c_aud"] for r in rows)
        lv     = sum(r["lv"]    for r in rows)
        mirs_r = sum(r["mirs_r"] for r in rows)
        mirs_a = sum(r["mirs_a"] for r in rows)
        wirs_r = sum(r["wirs_r"] for r in rows)
        wirs_a = sum(r["wirs_a"] for r in rows)
        ncrs_r = sum(r["ncrs_r"] for r in rows)
        ncrs_c = sum(r["ncrs_c"] for r in rows)

        sub_rate = round(ca     / recv   * 100, 1) if recv   else 0
        aud_rate = round(c_aud  / p_aud  * 100, 1) if p_aud  else 0
        mir_rate = round(mirs_a / mirs_r * 100, 1) if mirs_r else 0
        wir_rate = round(wirs_a / wirs_r * 100, 1) if wirs_r else 0
        ncr_rate = round(ncrs_c / ncrs_r * 100, 1) if ncrs_r else 0
        # Simple equal-weight composite (PIF to configure final weights)
        q_score  = round((sub_rate + aud_rate + mir_rate + wir_rate + ncr_rate) / 5, 2)

        wq.append([
            d["name"], MONTH_YEAR,
            recv, ca, cb, cc, sub_rate,
            p_aud, c_aud, aud_rate,
            lv,
            mirs_r, mirs_a, mir_rate,
            wirs_r, wirs_a, wir_rate,
            ncrs_r, ncrs_c, ncr_rate,
            q_score,
        ])

    out = OUTPUT_DIR / f"MASTER KPI Consolidation - {MONTH_YEAR}.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> None:
    if not SAFETY_TPL.exists():
        raise FileNotFoundError(f"Safety template not found: {SAFETY_TPL}")
    if not QUALITY_TPL.exists():
        raise FileNotFoundError(f"Quality template not found: {QUALITY_TPL}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Generating Safety KPI submissions...")
    for d in DEVCOS:
        out = OUTPUT_DIR / f"{d['short']} - Safety KPI - {MONTH_YEAR}.xlsx"
        fill_safety(d, out)
        print(f"  ✓ {out.name}")

    print("\nGenerating Quality KPI submissions...")
    for d in DEVCOS:
        out = OUTPUT_DIR / f"{d['short']} - Quality KPI - {MONTH_YEAR}.xlsx"
        fill_quality(d, out)
        print(f"  ✓ {out.name}")

    print("\nGenerating Master KPI Consolidation reference sheet...")
    make_master()

    print(f"\nAll files written to:\n  {OUTPUT_DIR}\n")

    print("DevCo profiles:")
    print(f"  {'Company':<38} {'Profile'}")
    print(f"  {'-'*38} {'-'*10}")
    for d in DEVCOS:
        print(f"  {d['name']:<38} {d['profile']}")

    print("""
Test data summary
─────────────────────────────────────────────────────────────────────────────
FMSCO       Excellent — 520k manhours, zero incidents, high near-miss/obs rates
AlMardoof   Good      — 280k manhours, 1 LTI, 1 recordable, solid quality rates
Enjaz       Average   — 175k manhours, 2 LTI, 3 recordable, lower quality rates
AFAC        Good      — 320k manhours, zero LTI, 1 recordable, good obs rates
─────────────────────────────────────────────────────────────────────────────
""")


if __name__ == "__main__":
    main()
