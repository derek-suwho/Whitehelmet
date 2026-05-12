"""Unit tests for FormulaExecutor."""

import io
import json

import openpyxl
import pytest


def make_xlsx(rows: list[dict]) -> bytes:
    """Build a simple xlsx from a list of row dicts keyed by column letter."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Write header row (row 1): column letters as headers
    col_letters = list(rows[0].keys())
    for i, letter in enumerate(col_letters, start=1):
        ws.cell(row=1, column=i, value=letter)

    # Write data rows starting at row 2
    col_index = {letter: i for i, letter in enumerate(col_letters, start=1)}
    for row_idx, row_data in enumerate(rows, start=2):
        for letter, value in row_data.items():
            ws.cell(row=row_idx, column=col_index[letter], value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_formula(target_column, formula_type, expression, weight=None, benchmark=None, scoring_rules=None):
    """Build a minimal TemplateFormula-like object."""

    class F:
        pass

    f = F()
    f.target_column = target_column
    f.formula_type = formula_type
    f.expression = expression
    f.weight = weight
    f.benchmark = benchmark
    f.scoring_rules = json.dumps(scoring_rules) if scoring_rules else None
    return f


def test_column_formula_divides_values():
    """=O{row}/N{row} should produce O/N for each data row."""
    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx(
        [
            {"N": 10.0, "O": 5.0},
            {"N": 20.0, "O": 4.0},
        ]
    )
    formulas = [make_formula("P", "column", "=O{row}/N{row}")]
    result = FormulaExecutor.execute(xlsx_bytes, formulas)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    p_col = headers["P"]
    assert ws.cell(row=2, column=p_col).value == pytest.approx(0.5)
    assert ws.cell(row=3, column=p_col).value == pytest.approx(0.2)


def test_single_cell_formula_sum():
    """=SUM(N2:N3) should produce the sum of column N."""
    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx([{"N": 10.0}, {"N": 20.0}])
    formulas = [make_formula("P", "single_cell", "=SUM(N2:N3)")]
    result = FormulaExecutor.execute(xlsx_bytes, formulas)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    p_col = headers["P"]
    assert ws.cell(row=2, column=p_col).value == pytest.approx(30.0)


def test_scoring_rules_applied():
    """Values should be mapped to scores based on scoring_rules."""
    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx([{"N": 10.0, "O": 1.0}])  # O/N = 0.1
    scoring = [
        {"min": None, "max": 0.05, "score": 100},
        {"min": 0.05, "max": 0.2, "score": 80},
        {"min": 0.2, "max": None, "score": 0},
    ]
    formulas = [make_formula("P", "column", "=O{row}/N{row}", scoring_rules=scoring)]
    result = FormulaExecutor.execute(xlsx_bytes, formulas)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    score_col = headers.get("P_score")
    assert score_col is not None
    assert ws.cell(row=2, column=score_col).value == 80


def test_no_formulas_returns_original():
    """If formulas list is empty, return original bytes unchanged."""
    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx([{"A": 1.0}])
    result = FormulaExecutor.execute(xlsx_bytes, [])
    assert result == xlsx_bytes


def test_weighted_score_summary():
    """Weighted total should be written to a summary cell when weights are present."""
    from app.services.formula_executor import FormulaExecutor

    scoring = [{"min": None, "max": 0.05, "score": 100}, {"min": 0.05, "max": None, "score": 0}]
    xlsx_bytes = make_xlsx([{"N": 10.0, "O": 0.2}])  # O/N = 0.02 → score 100
    formulas = [make_formula("P", "column", "=O{row}/N{row}", weight=0.5, scoring_rules=scoring)]
    result = FormulaExecutor.execute(xlsx_bytes, formulas)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Weighted Score":
                found = True
    assert found


def test_apply_scoring_none_value():
    """_apply_scoring returns None when value is None."""
    from app.services.formula_executor import _apply_scoring

    assert _apply_scoring(None, [{"min": 0, "max": 100, "score": 50}]) is None


def test_apply_scoring_no_match():
    """_apply_scoring returns None when no rule matches."""
    from app.services.formula_executor import _apply_scoring

    assert _apply_scoring(50.0, [{"min": 0, "max": 10, "score": 100}]) is None


def test_formula_targets_existing_column():
    """Formula targeting a column that already exists in headers reuses that column."""
    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx([{"N": 10.0, "O": 5.0}])
    formulas = [make_formula("N", "column", "=O{row}*2")]
    result = FormulaExecutor.execute(xlsx_bytes, formulas)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=2, column=headers["N"]).value == pytest.approx(10.0)


def test_xlcalculator_parse_failure():
    """When xlcalculator fails to parse, returns the file with formula strings intact."""
    from unittest.mock import patch

    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx([{"N": 10.0, "O": 5.0}])
    formulas = [make_formula("P", "column", "=O{row}/N{row}")]
    with patch("app.services.formula_executor.ModelCompiler") as MockCompiler:
        MockCompiler.return_value.read_and_parse_archive.side_effect = Exception("parse failed")
        result = FormulaExecutor.execute(xlsx_bytes, formulas)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert wb is not None


def test_scoring_column_reuse():
    """Score column created on first row is reused (not duplicated) on subsequent rows."""
    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx(
        [
            {"N": 10.0, "O": 1.0},  # O/N = 0.1 → score 80
            {"N": 10.0, "O": 0.3},  # O/N = 0.03 → score 100
        ]
    )
    scoring = [
        {"min": None, "max": 0.05, "score": 100},
        {"min": 0.05, "max": 0.2, "score": 80},
        {"min": 0.2, "max": None, "score": 0},
    ]
    formulas = [make_formula("P", "column", "=O{row}/N{row}", scoring_rules=scoring)]
    result = FormulaExecutor.execute(xlsx_bytes, formulas)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    score_col = headers.get("P_score")
    assert score_col is not None
    assert ws.cell(row=2, column=score_col).value == 80
    assert ws.cell(row=3, column=score_col).value == 100


def test_evaluator_exception_column():
    """Evaluator exception on column formula is caught; function returns valid xlsx."""
    from unittest.mock import patch

    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx([{"N": 10.0, "O": 5.0}])
    formulas = [make_formula("P", "column", "=O{row}/N{row}")]
    with (
        patch("app.services.formula_executor.ModelCompiler") as MockCompiler,
        patch("app.services.formula_executor.Evaluator") as MockEval,
    ):
        MockCompiler.return_value.read_and_parse_archive.return_value = object()
        MockEval.return_value.evaluate.side_effect = Exception("eval failed")
        result = FormulaExecutor.execute(xlsx_bytes, formulas)
    # Function completes without re-raising; result is a loadable xlsx with the P column
    ws = openpyxl.load_workbook(io.BytesIO(result)).active
    assert "P" in {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}


def test_evaluator_exception_single_cell():
    """Evaluator exception on single_cell formula is caught; function returns valid xlsx."""
    from unittest.mock import patch

    from app.services.formula_executor import FormulaExecutor

    xlsx_bytes = make_xlsx([{"N": 10.0}, {"N": 20.0}])
    formulas = [make_formula("P", "single_cell", "=SUM(N2:N3)")]
    with (
        patch("app.services.formula_executor.ModelCompiler") as MockCompiler,
        patch("app.services.formula_executor.Evaluator") as MockEval,
    ):
        MockCompiler.return_value.read_and_parse_archive.return_value = object()
        MockEval.return_value.evaluate.side_effect = Exception("eval failed")
        result = FormulaExecutor.execute(xlsx_bytes, formulas)
    ws = openpyxl.load_workbook(io.BytesIO(result)).active
    assert "P" in {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
