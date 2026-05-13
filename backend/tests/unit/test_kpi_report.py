"""Tests for KPI report generation — scoring functions + report builder."""

import io
import openpyxl
import pytest

from app.services.kpi_report import (
    _num,
    _rate,
    _score_fatality,
    _score_leadership,
    _score_ltifr,
    _score_near_miss,
    _score_rewards,
    _score_safety_obs,
    _score_trir,
    generate_safety_kpi_report,
)


# ── _score_fatality ──────────────────────────────────────────────────────────

def test_score_fatality_none():
    assert _score_fatality(None) is None

def test_score_fatality_zero():
    assert _score_fatality(0) == 1.0

def test_score_fatality_low():
    assert _score_fatality(0.005) == 0.9

def test_score_fatality_mid_low():
    assert _score_fatality(0.03) == 0.8

def test_score_fatality_mid():
    assert _score_fatality(0.08) == 0.7

def test_score_fatality_high():
    assert _score_fatality(0.5) == 0.0


# ── _score_ltifr ─────────────────────────────────────────────────────────────

def test_score_ltifr_none():
    assert _score_ltifr(None) is None

def test_score_ltifr_zero():
    assert _score_ltifr(0) == 1.0

def test_score_ltifr_low():
    assert _score_ltifr(0.02) == 0.9

def test_score_ltifr_mid():
    assert _score_ltifr(0.07) == 0.8

def test_score_ltifr_mid_high():
    assert _score_ltifr(0.15) == 0.7

def test_score_ltifr_high():
    assert _score_ltifr(0.5) == 0.0


# ── _score_trir ──────────────────────────────────────────────────────────────

def test_score_trir_none():
    assert _score_trir(None) is None

def test_score_trir_zero():
    assert _score_trir(0) == 1.0

def test_score_trir_low():
    assert _score_trir(0.03) == 0.9

def test_score_trir_mid():
    assert _score_trir(0.07) == 0.8

def test_score_trir_mid_high():
    assert _score_trir(0.12) == 0.7

def test_score_trir_high():
    assert _score_trir(0.3) == 0.0


# ── _score_near_miss ─────────────────────────────────────────────────────────

def test_score_near_miss_none():
    assert _score_near_miss(None) is None

def test_score_near_miss_zero():
    assert _score_near_miss(0) == 0.0

def test_score_near_miss_high():
    assert _score_near_miss(50) == 1.0

def test_score_near_miss_mid_high():
    assert _score_near_miss(35) == 0.9

def test_score_near_miss_mid():
    assert _score_near_miss(25) == 0.8

def test_score_near_miss_low():
    assert _score_near_miss(10) == 0.7


# ── _score_safety_obs ────────────────────────────────────────────────────────

def test_score_safety_obs_none():
    assert _score_safety_obs(None) is None

def test_score_safety_obs_high():
    assert _score_safety_obs(120) == 1.0

def test_score_safety_obs_good():
    assert _score_safety_obs(85) == 0.9

def test_score_safety_obs_mid():
    assert _score_safety_obs(65) == 0.8

def test_score_safety_obs_low():
    assert _score_safety_obs(45) == 0.7

def test_score_safety_obs_zero():
    assert _score_safety_obs(10) == 0.0


# ── _score_leadership ────────────────────────────────────────────────────────

def test_score_leadership_none():
    assert _score_leadership(None) is None

def test_score_leadership_zero():
    assert _score_leadership(0) == 0.0

def test_score_leadership_high():
    assert _score_leadership(10) == 1.0

def test_score_leadership_good():
    assert _score_leadership(6) == 0.95

def test_score_leadership_mid():
    assert _score_leadership(4) == 0.90

def test_score_leadership_low():
    assert _score_leadership(2.5) == 0.85

def test_score_leadership_very_low():
    assert _score_leadership(1) == 0.80


# ── _score_rewards ───────────────────────────────────────────────────────────

def test_score_rewards_none():
    assert _score_rewards(None) is None

def test_score_rewards_zero():
    assert _score_rewards(0) == 0.0

def test_score_rewards_high():
    assert _score_rewards(0.5) == 1.0

def test_score_rewards_good():
    assert _score_rewards(0.35) == 0.95

def test_score_rewards_mid():
    assert _score_rewards(0.25) == 0.90

def test_score_rewards_low():
    assert _score_rewards(0.15) == 0.85

def test_score_rewards_very_low():
    assert _score_rewards(0.05) == 0.80


# ── _num helper ──────────────────────────────────────────────────────────────

def test_num_valid():
    assert _num(42) == 42.0

def test_num_string():
    assert _num("3.14") == pytest.approx(3.14)

def test_num_none():
    assert _num(None) is None

def test_num_text():
    assert _num("abc") is None

def test_num_nan():
    assert _num(float("nan")) is None


# ── _rate helper ─────────────────────────────────────────────────────────────

def test_rate_normal():
    assert _rate(2, 100000) == pytest.approx(4.0)

def test_rate_none_metric():
    assert _rate(None, 100000) is None

def test_rate_zero_manhours():
    assert _rate(5, 0) is None

def test_rate_none_manhours():
    assert _rate(5, None) is None


# ── generate_safety_kpi_report ───────────────────────────────────────────────

def _make_ard_xlsx(tmp_path):
    """Create a minimal ARD sheet with data and a Total row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARD"
    # Row 9: headers (not used by report but realistic)
    ws["E9"] = "Description"
    ws["F9"] = "Total Manhours"
    # Row 10-11: data rows
    ws["F10"] = 100000
    ws["F11"] = 200000
    ws["H10"] = 0
    ws["H11"] = 1
    ws["I10"] = 2
    ws["I11"] = 3
    ws["K10"] = 1
    ws["K11"] = 2
    ws["L10"] = 50
    ws["L11"] = 60
    ws["M10"] = 100
    ws["M11"] = 120
    ws["N10"] = 5
    ws["N11"] = 8
    ws["O10"] = 10
    ws["O11"] = 15
    # Row 16: Total row (column E has "Total")
    ws["E16"] = "Total"
    ws["F16"] = 300000
    ws["H16"] = 1
    ws["I16"] = 5
    ws["K16"] = 3
    ws["L16"] = 110
    ws["M16"] = 220
    ws["N16"] = 13
    ws["O16"] = 25
    path = tmp_path / "test_ard.xlsx"
    wb.save(str(path))
    return str(path)


def test_generate_report_returns_bytes(tmp_path):
    path = _make_ard_xlsx(tmp_path)
    result = generate_safety_kpi_report(path, "TestDevCo")
    assert isinstance(result, bytes)
    assert len(result) > 100


def test_generate_report_has_kpi_sheet(tmp_path):
    path = _make_ard_xlsx(tmp_path)
    result = generate_safety_kpi_report(path, "TestDevCo")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert wb.active.title == "KPI Safety Report"


def test_generate_report_project_name(tmp_path):
    path = _make_ard_xlsx(tmp_path)
    result = generate_safety_kpi_report(path, "MyCompany")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert wb.active["B6"].value == "MyCompany"


def test_generate_report_has_scores(tmp_path):
    path = _make_ard_xlsx(tmp_path)
    result = generate_safety_kpi_report(path, "Test")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    # D6 = fatality score, should be a percentage string
    assert ws["D6"].value is not None
    assert "%" in str(ws["D6"].value)


def test_generate_report_raw_inputs(tmp_path):
    path = _make_ard_xlsx(tmp_path)
    result = generate_safety_kpi_report(path, "Test")
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert "300000" in str(ws["C8"].value)


def test_generate_report_no_total_row_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARD"
    ws["F10"] = 100
    path = tmp_path / "no_total.xlsx"
    wb.save(str(path))
    with pytest.raises(ValueError, match="Total row"):
        generate_safety_kpi_report(str(path))


def test_generate_report_uses_active_sheet_if_no_ard(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["E12"] = "Total"
    ws["F12"] = 50000
    ws["H12"] = 0
    ws["I12"] = 0
    ws["K12"] = 0
    ws["L12"] = 0
    ws["M12"] = 0
    ws["N12"] = 0
    ws["O12"] = 0
    path = tmp_path / "no_ard.xlsx"
    wb.save(str(path))
    result = generate_safety_kpi_report(str(path))
    assert isinstance(result, bytes)


def test_generate_report_total_in_col_a(tmp_path):
    """Total keyword in column A instead of column E."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARD"
    ws["A15"] = "Total"
    ws["F15"] = 100000
    ws["H15"] = 0
    ws["I15"] = 0
    ws["K15"] = 0
    ws["L15"] = 0
    ws["M15"] = 0
    ws["N15"] = 0
    ws["O15"] = 0
    path = tmp_path / "total_col_a.xlsx"
    wb.save(str(path))
    result = generate_safety_kpi_report(str(path))
    assert isinstance(result, bytes)
