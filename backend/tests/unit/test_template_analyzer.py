"""Test template layout analysis."""
import io
import openpyxl
from app.services.template_analyzer import analyze_template


def _make_safety_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARD"
    for col, header in [("F", "Manhours"), ("G", "Safe Manhours"), ("H", "Fatalities")]:
        ws[f"{col}9"] = header
    for row in range(10, 15):
        ws[f"F{row}"] = 100000
        ws[f"H{row}"] = row - 10
    ws["F16"] = "=SUM(F10:F14)"
    ws["H16"] = "=SUM(H10:H14)"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_analyze_detects_headers():
    result = analyze_template(_make_safety_template())
    assert len(result["sheets"]) == 1
    ard = result["sheets"][0]
    assert ard["name"] == "ARD"
    header_rows = [h for h in ard["headers"] if h["likely_header"]]
    assert any(h["row"] == 9 for h in header_rows)


def test_analyze_finds_existing_formulas():
    result = analyze_template(_make_safety_template())
    ard = result["sheets"][0]
    formula_cells = [f["cell"] for f in ard["existing_formulas"]]
    assert "F16" in formula_cells
    assert "H16" in formula_cells
