"""Tests for formula-aware consolidation: header_row, evaluate=False, remap methods."""

import io
import json

import openpyxl
import pytest


def make_xlsx(rows: list[dict], header_row: int = 1) -> bytes:
    """Build xlsx with headers at specified row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    col_letters = list(rows[0].keys())
    for i, letter in enumerate(col_letters, start=1):
        ws.cell(row=header_row, column=i, value=letter)
    col_index = {letter: i for i, letter in enumerate(col_letters, start=1)}
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        for letter, value in row_data.items():
            ws.cell(row=row_idx, column=col_index[letter], value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_formula(target_column, formula_type, expression, **kwargs):
    class F:
        pass
    f = F()
    f.target_column = target_column
    f.formula_type = formula_type
    f.expression = expression
    f.weight = kwargs.get("weight")
    f.benchmark = kwargs.get("benchmark")
    f.scoring_rules = json.dumps(kwargs["scoring_rules"]) if "scoring_rules" in kwargs else None
    f.target_row = kwargs.get("target_row")
    f.target_sheet = kwargs.get("target_sheet")
    return f


# ── header_row tests ─────────────────────────────────────────────


class TestHeaderRow:
    def test_header_row_9(self):
        """FormulaExecutor reads headers from row 9 when header_row=9."""
        from app.services.formula_executor import FormulaExecutor

        xlsx = make_xlsx([{"N": 10.0, "O": 5.0}], header_row=9)
        formulas = [make_formula("P", "column", "=O{row}/N{row}")]
        result = FormulaExecutor.execute(xlsx, formulas, header_row=9)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = {ws.cell(row=9, column=c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row=9, column=c).value}
        assert "P" in headers
        assert ws.cell(row=10, column=headers["P"]).value == pytest.approx(0.5)

    def test_header_row_default(self):
        """Default header_row=1 preserves existing behavior."""
        from app.services.formula_executor import FormulaExecutor

        xlsx = make_xlsx([{"N": 10.0, "O": 5.0}])
        formulas = [make_formula("P", "column", "=O{row}/N{row}")]
        result = FormulaExecutor.execute(xlsx, formulas, header_row=1)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.cell(row=2, column=3).value == pytest.approx(0.5)

    def test_data_start_follows_header_row(self):
        """Data rows start at header_row + 1."""
        from app.services.formula_executor import FormulaExecutor

        xlsx = make_xlsx([{"A": 1.0, "B": 2.0}, {"A": 3.0, "B": 4.0}], header_row=5)
        formulas = [make_formula("C", "column", "=A{row}+B{row}")]
        result = FormulaExecutor.execute(xlsx, formulas, header_row=5)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        # Data rows 6 and 7
        headers = {ws.cell(row=5, column=c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row=5, column=c).value}
        c_col = headers["C"]
        assert ws.cell(row=6, column=c_col).value == pytest.approx(3.0)
        assert ws.cell(row=7, column=c_col).value == pytest.approx(7.0)


# ── evaluate=False tests ─────────────────────────────────────────


class TestEvaluateFalse:
    def test_returns_formula_strings(self):
        """evaluate=False writes formula strings, not computed values."""
        from app.services.formula_executor import FormulaExecutor

        xlsx = make_xlsx([{"N": 10.0, "O": 5.0}])
        formulas = [make_formula("P", "column", "=O{row}/N{row}")]
        result = FormulaExecutor.execute(xlsx, formulas, evaluate=False)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
        p_col = headers["P"]
        cell = ws.cell(row=2, column=p_col)
        # Should be formula string, not numeric
        assert cell.value is not None
        assert str(cell.value).startswith("=")

    def test_no_scoring_columns(self):
        """evaluate=False should not produce scoring or weighted summary columns."""
        from app.services.formula_executor import FormulaExecutor

        scoring = [{"min": None, "max": 0.05, "score": 100}, {"min": 0.05, "max": None, "score": 0}]
        xlsx = make_xlsx([{"N": 10.0, "O": 0.2}])
        formulas = [make_formula("P", "column", "=O{row}/N{row}", weight=0.5, scoring_rules=scoring)]
        result = FormulaExecutor.execute(xlsx, formulas, evaluate=False)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        all_headers = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
        assert "P_score" not in all_headers
        assert "Weighted Score" not in {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}


# ── remap_expression tests ───────────────────────────────────────


class TestRemapExpression:
    def test_basic_column_remap(self):
        """Column letters remapped from template to consolidated positions."""
        from app.services.formula_executor import FormulaExecutor

        template_h2c = {"LTI": 1, "Manhours": 2}  # A, B
        target_h2c = {"Source File": 1, "LTI": 2, "Manhours": 3}  # A, B, C
        result = FormulaExecutor.remap_expression("=A{row}/B{row}", template_h2c, target_h2c)
        assert result == "=B{row}/C{row}"

    def test_rename_map(self):
        """AI renames handled via header_rename_map."""
        from app.services.formula_executor import FormulaExecutor

        template_h2c = {"Lost Time Injuries": 1}  # col A
        target_h2c = {"LTI Count": 2}  # col B in consolidated
        rename = {"Lost Time Injuries": "LTI Count"}
        result = FormulaExecutor.remap_expression("=A{row}*200000", template_h2c, target_h2c, rename)
        assert result == "=B{row}*200000"

    def test_unresolvable_passthrough(self):
        """Unresolvable references pass through unchanged."""
        from app.services.formula_executor import FormulaExecutor

        result = FormulaExecutor.remap_expression("=Z{row}+1", {}, {})
        assert result == "=Z{row}+1"

    def test_numeric_row_refs(self):
        """Expressions with numeric rows (e.g. =SUM(A2:A10)) also get remapped."""
        from app.services.formula_executor import FormulaExecutor

        template_h2c = {"Total": 1}  # col A
        target_h2c = {"Total": 3}  # col C
        result = FormulaExecutor.remap_expression("=SUM(A2:A10)", template_h2c, target_h2c)
        assert result == "=SUM(C2:C10)"


# ── remap_target_column tests ────────────────────────────────────


class TestRemapTargetColumn:
    def test_direct_match(self):
        """Target found directly in consolidated headers."""
        from app.services.formula_executor import FormulaExecutor

        result = FormulaExecutor.remap_target_column("LTI", {}, {"LTI": 3})
        assert result == "LTI"

    def test_case_insensitive(self):
        """Case-insensitive match."""
        from app.services.formula_executor import FormulaExecutor

        result = FormulaExecutor.remap_target_column("lti", {}, {"LTI": 3})
        assert result == "LTI"

    def test_derived_column_passthrough(self):
        """Derived columns (not in template or target) return as-is for appending."""
        from app.services.formula_executor import FormulaExecutor

        result = FormulaExecutor.remap_target_column("LTIFR", {"A": 1}, {"B": 2})
        assert result == "LTIFR"

    def test_rename_map_resolution(self):
        """AI rename resolves through rename map."""
        from app.services.formula_executor import FormulaExecutor

        rename = {"Lost Time Injuries": "LTI Count"}
        result = FormulaExecutor.remap_target_column(
            "Lost Time Injuries", {}, {"LTI Count": 5}, rename
        )
        assert result == "LTI Count"


# ── _detect_header_row tests ─────────────────────────────────────


class TestDetectHeaderRow:
    def test_headers_on_row_1(self):
        """Standard xlsx with headers on row 1."""
        from app.api.routes.templates import _detect_header_row

        wb = openpyxl.Workbook()
        ws = wb.active
        for c, h in enumerate(["Name", "Age", "City"], start=1):
            ws.cell(row=1, column=c, value=h)
        ws.cell(row=2, column=1, value=42)
        assert _detect_header_row(wb) == 1

    def test_headers_on_row_9(self):
        """Template with title rows before headers at row 9."""
        from app.api.routes.templates import _detect_header_row

        wb = openpyxl.Workbook()
        ws = wb.active
        # Sparse title rows
        ws.cell(row=1, column=1, value="Company Report")
        ws.cell(row=3, column=1, value="Period: Q1 2026")
        # Dense string headers at row 9
        for c, h in enumerate(["KPI", "Target", "Actual", "Score", "Weight"], start=1):
            ws.cell(row=9, column=c, value=h)
        assert _detect_header_row(wb) == 9

    def test_empty_sheet_returns_1(self):
        """Empty sheet defaults to row 1."""
        from app.api.routes.templates import _detect_header_row

        wb = openpyxl.Workbook()
        assert _detect_header_row(wb) == 1
