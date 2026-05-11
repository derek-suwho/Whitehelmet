"""Test FormulaExecutor with absolute cell references (Safety KPI style)."""
import io
import openpyxl
import pytest
from app.services.formula_executor import FormulaExecutor


class FakeFormula:
    def __init__(self, name, target_column, formula_type, expression,
                 target_row=None, target_sheet=None, weight=None, scoring_rules=None):
        self.name = name
        self.target_column = target_column
        self.formula_type = formula_type
        self.expression = expression
        self.target_row = target_row
        self.target_sheet = target_sheet
        self.weight = weight
        self.benchmark = None
        self.scoring_rules = scoring_rules


def _make_safety_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARD"
    ws["F9"] = "Total No. of Manhours"
    ws["H9"] = "No. of Fatality in the project"
    ws["I9"] = "Lost Time Injury incidents"
    ws["F10"] = 100000
    ws["F11"] = 200000
    ws["H10"] = 0
    ws["H11"] = 1
    ws["I10"] = 2
    ws["I11"] = 3
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_absolute_ref_sum_formula():
    xlsx = _make_safety_xlsx()
    formulas = [
        FakeFormula(
            name="Total Manhours",
            target_column="F",
            formula_type="single_cell",
            expression="=SUM(F10:F14)",
            target_row=16,
        ),
    ]
    result = FormulaExecutor.execute(xlsx, formulas)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    val = ws["F16"].value
    assert val == 300000, f"Expected 300000, got {val}"


def test_absolute_ref_kpi_formula_end_to_end():
    xlsx = _make_safety_xlsx()
    formulas = [
        FakeFormula(name="Total Manhours", target_column="F", formula_type="single_cell",
                    expression="=SUM(F10:F14)", target_row=16),
        FakeFormula(name="Total Fatalities", target_column="H", formula_type="single_cell",
                    expression="=SUM(H10:H14)", target_row=16),
        FakeFormula(name="Total LTI", target_column="I", formula_type="single_cell",
                    expression="=SUM(I10:I14)", target_row=16),
        FakeFormula(name="Fatality Rate", target_column="H", formula_type="single_cell",
                    expression="=H16*200000/F16", target_row=17, weight=0.20),
        FakeFormula(name="LTIFR", target_column="I", formula_type="single_cell",
                    expression="=I16*200000/F16", target_row=17, weight=0.10),
    ]
    result = FormulaExecutor.execute(xlsx, formulas)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws["F16"].value == 300000
    assert ws["H16"].value == 1
    assert ws["I16"].value == 5
    fr = ws["H17"].value
    assert fr is not None and abs(fr - (1 * 200000 / 300000)) < 0.001
    ltifr = ws["I17"].value
    assert ltifr is not None and abs(ltifr - (5 * 200000 / 300000)) < 0.001


def test_no_row_placeholder_passthrough():
    xlsx = _make_safety_xlsx()
    formulas = [
        FakeFormula(name="Test", target_column="Z", formula_type="single_cell",
                    expression="=H16*200000/F16", target_row=17),
    ]
    # Just needs the SUM formulas to work, so add them first
    formulas_with_deps = [
        FakeFormula(name="Total Manhours", target_column="F", formula_type="single_cell",
                    expression="=SUM(F10:F14)", target_row=16),
        FakeFormula(name="Total Fatalities", target_column="H", formula_type="single_cell",
                    expression="=SUM(H10:H14)", target_row=16),
    ] + formulas
    result = FormulaExecutor.execute(xlsx, formulas_with_deps)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws["Z17"].value is not None


def test_target_sheet_targeting():
    """Formulas with target_sheet write to the correct sheet."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1["A1"] = 10
    ws1["A2"] = 20
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "placeholder"
    buf = io.BytesIO()
    wb.save(buf)
    xlsx = buf.getvalue()

    formulas = [
        FakeFormula(name="Total", target_column="A", formula_type="single_cell",
                    expression="=SUM(Data!A1:A2)", target_row=2, target_sheet="Summary"),
    ]
    result = FormulaExecutor.execute(xlsx, formulas)
    wb2 = openpyxl.load_workbook(io.BytesIO(result))
    # Summary!A2 should have 30
    val = wb2["Summary"]["A2"].value
    assert val == 30, f"Expected 30, got {val}"
