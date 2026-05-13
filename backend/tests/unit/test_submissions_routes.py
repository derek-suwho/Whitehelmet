"""Tests for admin submission routes — covers get/download/review/patch/recalculate."""

import io
import uuid

import openpyxl
import pytest
from unittest.mock import patch, MagicMock

from app.core.dependencies import get_current_user, verify_csrf
from app.core.rbac import require_admin
from app.models.consolidated_sheet import ConsolidatedSheet
from app.models.organization import Organization
from app.models.submission import Submission
from app.models.template_assignment import TemplateAssignment
from app.models.template_formula import TemplateFormula


def _make_xlsx(path):
    """Create a minimal valid xlsx at path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARD"
    ws["A1"] = "Header"
    ws["A2"] = 100
    # Add Total row for KPI report
    ws["E16"] = "Total"
    ws["F16"] = 100000
    ws["H16"] = 0
    ws["I16"] = 0
    ws["K16"] = 0
    ws["L16"] = 0
    ws["M16"] = 0
    ws["N16"] = 0
    ws["O16"] = 0
    wb.save(str(path))
    return str(path)


def _create_submission(db, tmp_path, **overrides):
    """Insert a Submission record with a real xlsx file on disk."""
    file_path = tmp_path / f"{uuid.uuid4()}.xlsx"
    _make_xlsx(file_path)
    sub = Submission(
        id=overrides.get("id", str(uuid.uuid4())),
        assignment_id=overrides.get("assignment_id", str(uuid.uuid4())),
        org_id=overrides.get("org_id", str(uuid.uuid4())),
        file_path=str(file_path),
        file_name=overrides.get("file_name", "test.xlsx"),
        status=overrides.get("status", "submitted"),
        file_revision=overrides.get("file_revision", 0),
        processed_file_path=overrides.get("processed_file_path", None),
    )
    db.add(sub)
    db.commit()
    db.refresh(sub)
    return sub


# ── GET /api/admin/submissions/{id} ─────────────────────────────────────────

def test_get_submission_found(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    resp = pif_admin_client.get(f"/api/admin/submissions/{sub.id}")
    assert resp.status_code == 200
    data = resp.json()
    assert data["id"] == sub.id
    assert data["file_name"] == "test.xlsx"
    assert data["has_processed"] is False
    assert data["file_revision"] == 0


def test_get_submission_not_found(pif_admin_client):
    resp = pif_admin_client.get("/api/admin/submissions/nonexistent")
    assert resp.status_code == 404


# ── GET /api/admin/submissions/{id}/download ─────────────────────────────────

def test_download_raw(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    resp = pif_admin_client.get(f"/api/admin/submissions/{sub.id}/download?type=raw")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers.get("content-type", "")


def test_download_processed_not_available(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    resp = pif_admin_client.get(f"/api/admin/submissions/{sub.id}/download?type=processed")
    assert resp.status_code == 404


def test_download_processed_available(pif_admin_client, db, tmp_path):
    proc_path = tmp_path / "processed.xlsx"
    _make_xlsx(proc_path)
    sub = _create_submission(db, tmp_path, processed_file_path=str(proc_path))
    resp = pif_admin_client.get(f"/api/admin/submissions/{sub.id}/download?type=processed")
    assert resp.status_code == 200


def test_download_not_found(pif_admin_client):
    resp = pif_admin_client.get("/api/admin/submissions/nope/download")
    assert resp.status_code == 404


def test_download_file_missing_on_disk(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    import os
    os.unlink(sub.file_path)
    resp = pif_admin_client.get(f"/api/admin/submissions/{sub.id}/download?type=raw")
    assert resp.status_code == 404


# ── POST /api/admin/submissions/{id}/review ──────────────────────────────────

def test_review_approved(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    resp = pif_admin_client.post(
        f"/api/admin/submissions/{sub.id}/review",
        json={"status": "approved", "comment": "Looks good"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["review_status"] == "approved"
    assert data["review_comment"] == "Looks good"


def test_review_changes_requested(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    resp = pif_admin_client.post(
        f"/api/admin/submissions/{sub.id}/review",
        json={"status": "changes_requested", "comment": "Fix row 5"},
    )
    assert resp.status_code == 200
    assert resp.json()["review_status"] == "changes_requested"


def test_review_invalid_status(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    resp = pif_admin_client.post(
        f"/api/admin/submissions/{sub.id}/review",
        json={"status": "rejected"},
    )
    assert resp.status_code == 422


def test_review_not_found(pif_admin_client):
    resp = pif_admin_client.post(
        "/api/admin/submissions/nonexistent/review",
        json={"status": "approved"},
    )
    assert resp.status_code == 404


# ── PUT /api/admin/submissions/{id}/file ─────────────────────────────────────

def test_update_file(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    new_xlsx = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Updated"
    wb.save(new_xlsx)
    new_xlsx.seek(0)
    resp = pif_admin_client.put(
        f"/api/admin/submissions/{sub.id}/file",
        files={"file": ("updated.xlsx", new_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    assert resp.json()["file_name"] == "updated.xlsx"


def test_update_file_not_found(pif_admin_client):
    new_xlsx = io.BytesIO(b"dummy")
    resp = pif_admin_client.put(
        "/api/admin/submissions/nonexistent/file",
        files={"file": ("f.xlsx", new_xlsx)},
    )
    assert resp.status_code == 404


# ── POST /api/admin/submissions/{id}/kpi-report ─────────────────────────────

def test_kpi_report_success(pif_admin_client, db, tmp_path):
    from app.core.config import Settings
    s = Settings(
        anthropic_api_key="", openrouter_api_key="",
        upload_dir=str(tmp_path / "uploads"), max_upload_size_mb=50,
        session_secret="test", csrf_secret="test", db_password="",
    )
    org = Organization(id=str(uuid.uuid4()), name="TestOrg", type="devco")
    db.add(org)
    db.commit()
    sub = _create_submission(db, tmp_path, org_id=org.id)
    with patch("app.api.routes.submissions.get_settings", return_value=s):
        resp = pif_admin_client.post(f"/api/admin/submissions/{sub.id}/kpi-report")
    assert resp.status_code == 200
    data = resp.json()
    assert "sheet_id" in data
    assert "TestOrg" in data["name"]


def test_kpi_report_not_found(pif_admin_client):
    resp = pif_admin_client.post("/api/admin/submissions/nope/kpi-report")
    assert resp.status_code == 404


def test_kpi_report_file_missing(pif_admin_client, db, tmp_path, settings_override):
    sub = _create_submission(db, tmp_path)
    import os
    os.unlink(sub.file_path)
    resp = pif_admin_client.post(f"/api/admin/submissions/{sub.id}/kpi-report")
    assert resp.status_code == 404


# ── PATCH /api/admin/submissions/{id}/cells ──────────────────────────────────

def test_patch_cells(pif_admin_client, db, tmp_path):
    proc_path = tmp_path / "proc.xlsx"
    _make_xlsx(proc_path)
    sub = _create_submission(db, tmp_path, processed_file_path=str(proc_path))
    resp = pif_admin_client.patch(
        f"/api/admin/submissions/{sub.id}/cells",
        json={"changes": [{"row": 1, "col": 1, "value": "Patched"}], "revision": 0},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["patches_applied"] == 1
    assert data["revision"] == 1


def test_patch_cells_no_processed_file(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path)
    resp = pif_admin_client.patch(
        f"/api/admin/submissions/{sub.id}/cells",
        json={"changes": [{"row": 1, "col": 1, "value": "X"}]},
    )
    assert resp.status_code == 400


def test_patch_cells_stale_revision(pif_admin_client, db, tmp_path):
    proc_path = tmp_path / "proc2.xlsx"
    _make_xlsx(proc_path)
    sub = _create_submission(db, tmp_path, processed_file_path=str(proc_path), file_revision=5)
    resp = pif_admin_client.patch(
        f"/api/admin/submissions/{sub.id}/cells",
        json={"changes": [{"row": 1, "col": 1, "value": "X"}], "revision": 3},
    )
    assert resp.status_code == 409


def test_patch_cells_not_found(pif_admin_client):
    resp = pif_admin_client.patch(
        "/api/admin/submissions/nope/cells",
        json={"changes": [{"row": 1, "col": 1, "value": "X"}]},
    )
    assert resp.status_code == 404


# ── POST /api/admin/submissions/{id}/recalculate ─────────────────────────────

def test_recalculate_success(pif_admin_client, db, tmp_path):
    tv_id = str(uuid.uuid4())
    assign = TemplateAssignment(
        id=str(uuid.uuid4()),
        template_version_id=tv_id,
        org_id=str(uuid.uuid4()),
    )
    db.add(assign)
    db.commit()
    sub = _create_submission(db, tmp_path, assignment_id=assign.id)
    formula = TemplateFormula(
        id=str(uuid.uuid4()),
        template_version_id=tv_id,
        name="Sum F",
        target_column="F",
        formula_type="single_cell",
        expression="=SUM(F10:F14)",
        target_row=16,
    )
    db.add(formula)
    db.commit()
    resp = pif_admin_client.post(f"/api/admin/submissions/{sub.id}/recalculate")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "recalculated"
    assert data["revision"] == 1


def test_recalculate_not_found(pif_admin_client):
    resp = pif_admin_client.post("/api/admin/submissions/nope/recalculate")
    assert resp.status_code == 404


def test_recalculate_no_assignment(pif_admin_client, db, tmp_path):
    sub = _create_submission(db, tmp_path, assignment_id="nonexistent")
    resp = pif_admin_client.post(f"/api/admin/submissions/{sub.id}/recalculate")
    assert resp.status_code == 404


def test_recalculate_no_formulas(pif_admin_client, db, tmp_path):
    tv_id = str(uuid.uuid4())
    assign = TemplateAssignment(
        id=str(uuid.uuid4()),
        template_version_id=tv_id,
        org_id=str(uuid.uuid4()),
    )
    db.add(assign)
    db.commit()
    sub = _create_submission(db, tmp_path, assignment_id=assign.id)
    resp = pif_admin_client.post(f"/api/admin/submissions/{sub.id}/recalculate")
    assert resp.status_code == 400
