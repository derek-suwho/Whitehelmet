"""KPI Report Generation — reads a DevCo's ARD sheet and produces a scored master sheet."""

from __future__ import annotations

import io
import uuid
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Scoring rules (from Sheet1 IFS formulas in KPI Template) ─────────────────

def _score_fatality(r: float | None) -> float | None:
    if r is None: return None
    if r == 0: return 1.0
    if r <= 0.01: return 0.9
    if r <= 0.05: return 0.8
    if r <= 0.1: return 0.7
    return 0.0

def _score_ltifr(r: float | None) -> float | None:
    if r is None: return None
    if r == 0: return 1.0
    if r <= 0.04: return 0.9
    if r <= 0.1: return 0.8
    if r <= 0.2: return 0.7
    return 0.0

def _score_trir(r: float | None) -> float | None:
    if r is None: return None
    if r == 0: return 1.0
    if r <= 0.045: return 0.9
    if r <= 0.09: return 0.8
    if r <= 0.15: return 0.7
    return 0.0

def _score_near_miss(r: float | None) -> float | None:
    if r is None: return None
    if r == 0: return 0.0
    if r > 40: return 1.0
    if r >= 30: return 0.9
    if r >= 20: return 0.8
    return 0.7

def _score_safety_obs(r: float | None) -> float | None:
    if r is None: return None
    if r >= 100: return 1.0
    if r >= 80: return 0.9
    if r >= 60: return 0.8
    if r >= 40: return 0.7
    return 0.0

def _score_leadership(r: float | None) -> float | None:
    if r is None: return None
    if r == 0: return 0.0
    if r > 7: return 1.0
    if r >= 5: return 0.95
    if r >= 3: return 0.90
    if r >= 2: return 0.85
    return 0.80

def _score_rewards(r: float | None) -> float | None:
    if r is None: return None
    if r == 0: return 0.0
    if r > 0.4: return 1.0
    if r >= 0.3: return 0.95
    if r >= 0.2: return 0.90
    if r >= 0.1: return 0.85
    return 0.80


def _num(v: Any) -> float | None:
    try:
        f = float(v)
        return f if f == f else None  # exclude NaN
    except (TypeError, ValueError):
        return None


def _rate(metric: float | None, manhours: float | None) -> float | None:
    if metric is None or not manhours:
        return None
    return metric * 200_000 / manhours


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_safety_kpi_report(
    submission_path: str,
    project_name: str = "DevCo",
) -> bytes:
    """
    Opens a DevCo's submitted ARD xlsx, computes KPI rates and scores,
    and returns a filled master report xlsx as bytes.
    """
    wb_in = openpyxl.load_workbook(submission_path, data_only=True)

    # Find the ARD / safety data sheet
    ard_sheet = None
    for name in wb_in.sheetnames:
        if name.upper() in ("ARD", "SAFETY") or "ARD" in name.upper():
            ard_sheet = wb_in[name]
            break
    if ard_sheet is None:
        ard_sheet = wb_in.active

    # Locate the Total row (column E or A contains "Total")
    total_row_idx = None
    for r in range(8, min(ard_sheet.max_row + 1, 30)):
        va = str(ard_sheet.cell(r, 1).value or "").strip().lower()
        ve = str(ard_sheet.cell(r, 5).value or "").strip().lower()
        if "total" in va or "total" in ve:
            total_row_idx = r
            break

    if total_row_idx is None:
        raise ValueError("Could not find Total row in submitted ARD sheet")

    # Extract column totals (F=6, H=8, I=9, K=11, L=12, M=13, N=14, O=15)
    def col(c: int) -> float | None:
        return _num(ard_sheet.cell(total_row_idx, c).value)

    manhours   = col(6)
    fatalities = col(8)
    lti        = col(9)
    trir_count = col(11)
    near_miss  = col(12)
    safety_obs = col(13)
    leadership = col(14)
    rewards    = col(15)

    # Compute KPI rates
    fatality_rate  = _rate(fatalities, manhours)
    ltifr          = _rate(lti, manhours)
    trir_rate      = _rate(trir_count, manhours)
    near_miss_rate = _rate(near_miss, manhours)
    safety_rate    = _rate(safety_obs, manhours)
    leadership_rate= _rate(leadership, manhours)
    rewards_rate   = _rate(rewards, manhours)

    # Apply scoring
    scores = {
        "fatality":   (_score_fatality(fatality_rate),  fatality_rate),
        "ltifr":      (_score_ltifr(ltifr),              ltifr),
        "trir":       (_score_trir(trir_rate),            trir_rate),
        "near_miss":  (_score_near_miss(near_miss_rate),  near_miss_rate),
        "safety_obs": (_score_safety_obs(safety_rate),    safety_rate),
        "leadership": (_score_leadership(leadership_rate),leadership_rate),
        "rewards":    (_score_rewards(rewards_rate),      rewards_rate),
    }

    # ── Build output workbook ────────────────────────────────────────────────
    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "KPI Safety Report"

    # Styling helpers
    header_fill  = PatternFill("solid", fgColor="FFF2CC")  # light yellow (matches template)
    green_fill   = PatternFill("solid", fgColor="E2EFDA")
    bold         = Font(bold=True)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def hdr(cell, value, fill=None, bold_=True):
        ws[cell] = value
        ws[cell].alignment = center
        if fill: ws[cell].fill = fill
        if bold_: ws[cell].font = bold

    # Row 1: title
    ws.merge_cells("A1:P1")
    hdr("A1", "Safety KPI Scoring Report", fill=PatternFill("solid", fgColor="1F4E79"))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)

    # Row 3: column group headers
    col_headers = [
        ("B3", "Project Name"),
        ("C3", "No. of Fatality in the project"),
        ("E3", "Lost Time Injury incidents"),
        ("G3", "Total Recordable Incident"),
        ("I3", "Total Near Miss incidents"),
        ("K3", "Total Safety Observations"),
        ("M3", "Total Leadership safety walks conducted"),
        ("O3", "Total number of Recognitions & rewards"),
    ]
    for addr, label in col_headers:
        hdr(addr, label, fill=header_fill)

    # Row 4: score criteria (brief)
    criteria = [
        ("C4", "0: 100%  >0.01: 90%  >0.05: 80%  >0.1: 70%  else: 0%"),
        ("E4", "0: 100%  ≤0.04: 90%  ≤0.1: 80%  ≤0.2: 70%  else: 0%"),
        ("G4", "0: 100%  ≤0.045: 90%  ≤0.09: 80%  ≤0.15: 70%  else: 0%"),
        ("I4", ">40: 100%  ≥30: 90%  ≥20: 80%  <20: 70%  0: 0%"),
        ("K4", "≥100: 100%  ≥80: 90%  ≥60: 80%  ≥40: 70%  <40: 0%"),
        ("M4", ">7: 100%  ≥5: 95%  ≥3: 90%  ≥2: 85%  <2: 80%  0: 0%"),
        ("O4", ">0.4: 100%  ≥0.3: 95%  ≥0.2: 90%  ≥0.1: 85%  <0.1: 80%  0: 0%"),
    ]
    for addr, label in criteria:
        ws[addr] = label
        ws[addr].alignment = Alignment(horizontal="center", wrap_text=True)
        ws[addr].font = Font(size=8, italic=True)
        ws[addr].fill = header_fill

    # Row 5: column short names
    short_names = [
        ("B5", "Project Name"),
        ("C5", "Rate"), ("D5", "Score"),
        ("E5", "Rate"), ("F5", "Score"),
        ("G5", "Rate"), ("H5", "Score"),
        ("I5", "Rate"), ("J5", "Score"),
        ("K5", "Rate"), ("L5", "Score"),
        ("M5", "Rate"), ("N5", "Score"),
        ("O5", "Rate"), ("P5", "Score"),
    ]
    for addr, label in short_names:
        hdr(addr, label, fill=header_fill)

    # Row 6: data
    def pct(v):
        if v is None: return "N/A"
        return f"{v:.1%}"

    def val_str(v):
        if v is None: return "N/A"
        return round(v, 4)

    ws["B6"] = project_name
    ws["B6"].font = bold

    data_cells = [
        ("C6", val_str(scores["fatality"][1])),   ("D6", pct(scores["fatality"][0])),
        ("E6", val_str(scores["ltifr"][1])),       ("F6", pct(scores["ltifr"][0])),
        ("G6", val_str(scores["trir"][1])),        ("H6", pct(scores["trir"][0])),
        ("I6", val_str(scores["near_miss"][1])),   ("J6", pct(scores["near_miss"][0])),
        ("K6", val_str(scores["safety_obs"][1])),  ("L6", pct(scores["safety_obs"][0])),
        ("M6", val_str(scores["leadership"][1])),  ("N6", pct(scores["leadership"][0])),
        ("O6", val_str(scores["rewards"][1])),     ("P6", pct(scores["rewards"][0])),
    ]
    for addr, value in data_cells:
        ws[addr] = value
        ws[addr].alignment = center
        # Colour score cells green/amber
        col_letter = addr[0]
        if col_letter in ("D", "F", "H", "J", "L", "N", "P") and isinstance(value, str) and "%" in value:
            pct_val = float(value.replace("%", "")) / 100
            ws[addr].fill = green_fill if pct_val >= 0.7 else PatternFill("solid", fgColor="FCE4D6")

    # Row 8: raw input summary
    ws["B8"] = "Raw Inputs"
    ws["B8"].font = bold
    ws["C8"] = f"Total Manhours: {manhours or 'N/A'}"
    ws["E8"] = f"Fatalities: {fatalities or 0}"
    ws["G8"] = f"LTI incidents: {lti or 0}"
    ws["I8"] = f"TRIR incidents: {trir_count or 0}"
    ws["K8"] = f"Near Miss: {near_miss or 0}"
    ws["M8"] = f"Safety Obs: {safety_obs or 0}"
    ws["O8"] = f"Leadership walks: {leadership or 0}"

    # Column widths
    ws.column_dimensions["B"].width = 28
    for c in ["C","D","E","F","G","H","I","J","K","L","M","N","O","P"]:
        ws.column_dimensions[c].width = 14

    ws.row_dimensions[3].height = 60
    ws.row_dimensions[4].height = 45

    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue()
