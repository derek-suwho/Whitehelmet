"""
Analyze template .xlsx files to extract layout metadata.
Used by the AI formula configuration flow to understand sheet structure
before generating TemplateFormula records.
"""
import io
import openpyxl
from openpyxl.utils import get_column_letter


def analyze_template(file_bytes: bytes) -> dict:
    """Extract structural metadata from an xlsx template.

    Returns dict with:
      - sheets: list of sheet info dicts
      - per sheet: headers (row, columns), data_range, existing_formulas, sample_data
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    result = {"sheets": []}

    for ws in wb.worksheets:
        sheet_info = {
            "name": ws.title,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "headers": [],
            "existing_formulas": [],
            "sample_data": [],
        }

        # Scan for header rows (rows where most cells are strings)
        for row_idx in range(1, min(ws.max_row + 1, 20)):
            row_cells = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_cells.append({
                        "col": get_column_letter(col_idx),
                        "value": str(cell.value)[:100],
                        "is_formula": str(cell.value).startswith("=") if cell.value else False,
                    })
            if row_cells:
                string_count = sum(1 for c in row_cells if not c["is_formula"] and not _is_numeric(c["value"]))
                sheet_info["headers"].append({
                    "row": row_idx,
                    "cells": row_cells,
                    "likely_header": string_count > len(row_cells) * 0.5,
                })

        # Collect existing formulas
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    sheet_info["existing_formulas"].append({
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "formula": str(cell.value),
                    })

        result["sheets"].append(sheet_info)

    return result


def _is_numeric(val: str) -> bool:
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False
