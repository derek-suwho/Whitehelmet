"""FormulaExecutor — applies Excel formulas to submitted xlsx files."""

from __future__ import annotations

import io
import json
import logging
import re
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string
from xlcalculator import Evaluator, ModelCompiler

logger = logging.getLogger(__name__)


def _col_index_to_letter(index: int) -> str:
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result


def _translate_formula_columns(expression: str, header_to_col: dict[str, int]) -> str:
    """Translate header names to actual Excel column letters in a formula.

    Formulas use header names as column references (e.g. =O{row}/N{row} where
    "O" and "N" are header names). This translates them to the actual column
    letters based on the header row position.

    For absolute cell references (e.g. =SUM(F10:F14)), if no header matches the
    column token, it passes through unchanged (name_to_letter.get falls back to
    the original name, which is already a valid column letter).
    """
    name_to_letter = {name: _col_index_to_letter(col_idx) for name, col_idx in header_to_col.items()}

    def replace(m: re.Match) -> str:
        col_name = m.group(1)
        row_part = m.group(2)
        return name_to_letter.get(col_name, col_name) + row_part

    return re.sub(r"([A-Z]+)(\{row\}|\d+)", replace, expression)


def _apply_scoring(value: float, rules: list[dict]) -> int | None:
    """Map a numeric value to a score using range rules. Rules checked in order."""
    if value is None:
        return None
    for rule in rules:
        lo = rule.get("min")
        hi = rule.get("max")
        if (lo is None or value >= lo) and (hi is None or value <= hi):
            return rule.get("score")
    return None


class FormulaExecutor:
    @staticmethod
    def execute(xlsx_bytes: bytes, formulas: list[Any], *, header_row: int = 1, evaluate: bool = True) -> bytes:
        """
        Apply formulas to an xlsx file and return modified bytes.

        formulas: list of TemplateFormula ORM objects (or duck-typed equivalents).
        Each has: target_column, formula_type, expression, weight, benchmark, scoring_rules (JSON string).
        Optional: target_row (int | None), target_sheet (str | None).

        Returns original bytes unchanged if formulas list is empty.
        """
        if not formulas:
            return xlsx_bytes

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

        # Group formulas by their target sheet (None → active sheet)
        def _resolve_sheet_name(formula: Any) -> str:
            ts = getattr(formula, "target_sheet", None)
            if ts:
                return ts
            return wb.active.title

        # Build per-sheet state: header_to_col, data_start, data_end, worksheet
        sheet_state: dict[str, dict] = {}

        def _get_sheet_state(sheet_name: str) -> dict:
            if sheet_name in sheet_state:
                return sheet_state[sheet_name]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            hr = header_row  # 1-indexed
            header_to_col: dict[str, int] = {}
            for col_idx in range(1, ws.max_column + 1):
                hdr = ws.cell(row=hr, column=col_idx).value
                if hdr:
                    header_to_col[str(hdr)] = col_idx
            state = {
                "ws": ws,
                "header_to_col": header_to_col,
                "data_start": hr + 1,
                "data_end": ws.max_row,
            }
            sheet_state[sheet_name] = state
            return state

        # Pre-populate state for all sheets referenced by formulas
        for formula in formulas:
            _get_sheet_state(_resolve_sheet_name(formula))

        # ── Write formula strings into cells ─────────────────────────────────
        # Map (sheet_name, target_column) → target_col_idx for later evaluation
        target_col_indices: dict[tuple[str, str], int] = {}

        for formula in formulas:
            sheet_name = _resolve_sheet_name(formula)
            state = _get_sheet_state(sheet_name)
            ws = state["ws"]
            header_to_col = state["header_to_col"]
            data_start = state["data_start"]
            data_end = state["data_end"]

            target = formula.target_column.upper()
            formula_target_row = getattr(formula, "target_row", None)

            if target in header_to_col:
                # Found as a named header in row 1
                target_col_idx = header_to_col[target]
            elif formula_target_row is not None and re.fullmatch(r"[A-Z]+", target):
                # Absolute-ref mode: target_row is set, treat target_column as Excel column letter
                target_col_idx = column_index_from_string(target)
                header_to_col[target] = target_col_idx
            else:
                # New column: append after last column and write header
                target_col_idx = ws.max_column + 1
                ws.cell(row=1, column=target_col_idx, value=target)
                header_to_col[target] = target_col_idx

            target_col_indices[(sheet_name, target)] = target_col_idx

            translated = _translate_formula_columns(formula.expression, header_to_col)

            if formula.formula_type == "column":
                for row in range(data_start, data_end + 1):
                    cell_formula = translated.replace("{row}", str(row))
                    ws.cell(row=row, column=target_col_idx, value=cell_formula)
            else:
                # single_cell: write formula once at target_row (or data_start if not set)
                write_row = getattr(formula, "target_row", None) or data_start
                ws.cell(row=write_row, column=target_col_idx, value=translated)

        if not evaluate:
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        # ── Evaluate with xlcalculator ────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        try:
            compiler = ModelCompiler()
            model = compiler.read_and_parse_archive(buf, ignore_sheets=[])
            evaluator = Evaluator(model)
        except Exception as exc:
            logger.warning("xlcalculator failed to parse workbook: %s — returning unprocessed file", exc)
            buf.seek(0)
            return buf.read()

        # ── Write computed values back (replace formula strings with numbers) ─
        weighted: list[tuple[float, float]] = []  # (weight, score)

        for formula in formulas:
            sheet_name = _resolve_sheet_name(formula)
            state = _get_sheet_state(sheet_name)
            ws = state["ws"]
            header_to_col = state["header_to_col"]
            data_start = state["data_start"]
            data_end = state["data_end"]

            target = formula.target_column.upper()
            target_col_idx = target_col_indices[(sheet_name, target)]
            scoring_rules = json.loads(formula.scoring_rules) if formula.scoring_rules else []

            if formula.formula_type == "column":
                formula_scores: list[float] = []
                for row in range(data_start, data_end + 1):
                    cell_addr = f"{sheet_name}!{_col_index_to_letter(target_col_idx)}{row}"
                    try:
                        computed = evaluator.evaluate(cell_addr)
                        numeric = float(computed) if computed not in (None, "") else None
                    except Exception:
                        numeric = None
                    ws.cell(row=row, column=target_col_idx, value=numeric)

                    if scoring_rules and numeric is not None:
                        score = _apply_scoring(numeric, scoring_rules)
                        score_header = f"{target}_score"
                        if score_header not in header_to_col:
                            score_col_idx = ws.max_column + 1
                            ws.cell(row=1, column=score_col_idx, value=score_header)
                            header_to_col[score_header] = score_col_idx
                        else:
                            score_col_idx = header_to_col[score_header]
                        ws.cell(row=row, column=score_col_idx, value=score)
                        formula_scores.append(score or 0)

                # Average scores across all rows, then apply weight once per formula
                if formula.weight and formula_scores:
                    avg_score = sum(formula_scores) / len(formula_scores)
                    weighted.append((formula.weight, avg_score))

            else:
                # single_cell: evaluate at the same row we wrote to
                write_row = getattr(formula, "target_row", None) or data_start
                cell_addr = f"{sheet_name}!{_col_index_to_letter(target_col_idx)}{write_row}"
                try:
                    computed = evaluator.evaluate(cell_addr)
                    numeric = float(computed) if computed not in (None, "") else None
                except Exception:
                    numeric = None
                ws.cell(row=write_row, column=target_col_idx, value=numeric)

        # ── Append weighted score summary (written to active sheet) ───────────
        if weighted:
            ws_active = wb.active
            total_weight = sum(w for w, _ in weighted)
            weighted_score = sum(w * s for w, s in weighted) / total_weight if total_weight > 0 else 0
            # Find the maximum row across all sheets used
            max_row = max(state["data_end"] for state in sheet_state.values())
            summary_row = max_row + 2
            ws_active.cell(row=summary_row, column=1, value="Weighted Score")
            ws_active.cell(row=summary_row, column=2, value=round(weighted_score, 2))

        result_buf = io.BytesIO()
        wb.save(result_buf)
        return result_buf.getvalue()

    @staticmethod
    def remap_expression(
        expression: str,
        template_header_to_col: dict[str, int],
        target_header_to_col: dict[str, int],
        header_rename_map: dict[str, str] | None = None,
    ) -> str:
        """Remap formula column references from template layout to target layout.

        Args:
            template_header_to_col: template header name → column index
            target_header_to_col: consolidated header name → column index
            header_rename_map: template header → unified header name (from AI column_map)
        """
        template_letter_to_name: dict[str, str] = {}
        for name, col_idx in template_header_to_col.items():
            letter = _col_index_to_letter(col_idx)
            template_letter_to_name[letter] = name

        rename = header_rename_map or {}
        name_to_target_letter: dict[str, str] = {
            name: _col_index_to_letter(col_idx)
            for name, col_idx in target_header_to_col.items()
        }

        def resolve_target_letter(template_name: str) -> str | None:
            if template_name in name_to_target_letter:
                return name_to_target_letter[template_name]
            unified_name = rename.get(template_name)
            if unified_name and unified_name in name_to_target_letter:
                return name_to_target_letter[unified_name]
            for src, dst in rename.items():
                if src.upper() == template_name.upper() and dst in name_to_target_letter:
                    return name_to_target_letter[dst]
            return None

        def replace(m: re.Match) -> str:
            col_ref = m.group(1)
            row_part = m.group(2)
            header_name = template_letter_to_name.get(col_ref)
            if header_name:
                target_letter = resolve_target_letter(header_name)
                if target_letter:
                    return target_letter + row_part
            target_letter = resolve_target_letter(col_ref)
            if target_letter:
                return target_letter + row_part
            return col_ref + row_part

        return re.sub(r"([A-Z]+)(\{row\}|\d+)", replace, expression)

    @staticmethod
    def remap_target_column(
        target: str,
        template_header_to_col: dict[str, int],
        target_header_to_col: dict[str, int],
        header_rename_map: dict[str, str] | None = None,
    ) -> str:
        """Remap a formula's target_column from template to target layout.
        Returns the header name in the target layout. For derived columns
        (e.g. LTIFR) that don't exist yet, returns the original target name
        so it can be appended as a new column."""
        rename = header_rename_map or {}

        target_upper = target.upper()
        for name in target_header_to_col:
            if name.upper() == target_upper:
                return name
        for src, dst in rename.items():
            if src.upper() == target_upper:
                for name in target_header_to_col:
                    if name.upper() == dst.upper():
                        return name
        template_letter_to_name = {
            _col_index_to_letter(idx): name
            for name, idx in template_header_to_col.items()
        }
        header_name = template_letter_to_name.get(target_upper)
        if header_name:
            renamed = rename.get(header_name, header_name)
            for name in target_header_to_col:
                if name.upper() == renamed.upper():
                    return name
            return renamed
        return target
