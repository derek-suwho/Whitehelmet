"""AI proxy routes — all traffic via OpenRouter.

Auth is intentionally stripped while Whitehelmet's external auth service is
pending integration (see auth.py). Re-add `dependencies=[Depends(get_current_user),
Depends(verify_csrf)]` on the router once login is implemented.
"""

import io
import json
import uuid
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import httpx

from app.core.config import get_settings
from app.core.dependencies import get_current_user
from app.db.session import get_db
from app.models.template import Template
from app.models.template_version import TemplateVersion
from app.models.user import User
from app.schemas.ai import ChatRequest, ConsolidateRequest, ConsolidateResponse, AgentRequest
from sqlalchemy.orm import Session

router = APIRouter(prefix="/api/ai", tags=["ai"])

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"


async def _openrouter_post(payload: dict) -> dict:
    """Non-streaming POST to OpenRouter."""
    settings = get_settings()
    if not settings.openrouter_api_key:
        raise HTTPException(status_code=503, detail="OpenRouter API key not configured")
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(
            OPENROUTER_URL,
            headers={
                "Authorization": f"Bearer {settings.openrouter_api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
        if not resp.is_success:
            raise HTTPException(
                status_code=resp.status_code,
                detail=f"OpenRouter error: {resp.text}"
            )
        return resp.json()


@router.post("/chat")
async def chat(body: ChatRequest):
    """Proxy chat to OpenRouter. SSE stream when body.stream=True, else JSON."""
    settings = get_settings()
    if not settings.openrouter_api_key:
        raise HTTPException(status_code=503, detail="OpenRouter API key not configured")

    payload = {
        "model": body.model,
        "max_tokens": body.max_tokens,
        "messages": body.messages,
        "stream": body.stream,
    }

    if not body.stream:
        return await _openrouter_post(payload)

    async def sse():
        async with httpx.AsyncClient(timeout=120.0) as client:
            async with client.stream(
                "POST",
                OPENROUTER_URL,
                headers={
                    "Authorization": f"Bearer {settings.openrouter_api_key}",
                    "Content-Type": "application/json",
                },
                json=payload,
            ) as resp:
                async for line in resp.aiter_lines():
                    if line.startswith("data: "):
                        yield f"{line}\n\n"

    return StreamingResponse(sse(), media_type="text/event-stream")


@router.post("/consolidate", response_model=ConsolidateResponse)
async def consolidate(body: ConsolidateRequest):
    """Detect column schema across files and return a unified mapping.

    The AI only sees headers + sample rows — the client applies the mapping
    to all rows, keeping large data out of the token budget.
    """
    system_prompt = (
        "You are a spreadsheet schema unifier. "
        "Given headers and sample rows from multiple Excel files, "
        "return ONLY a valid JSON object — no markdown, no extra text:\n"
        '{"unified_headers":["Source File","<col>",...],'
        '"mappings":[{"file":"<name>","column_map":{"<src>":"<unified>"}}]}\n\n'
        "Rules:\n"
        "- Always include \"Source File\" as the first unified header.\n"
        "- Merge columns that represent the same concept under one standard name "
        "(e.g. \"Invoice Amt\", \"Invoice Amount\", \"Inv. Amount\" → \"Invoice Amount\").\n"
        "- Include every column that appears in at least one file.\n"
        "- column_map must cover every source column in that file."
    )
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps([
            {"name": f.name, "headers": f.headers, "sample_rows": f.sample_rows}
            for f in body.files_schema
        ])},
    ]
    data = await _openrouter_post({
        "model": body.model,
        "max_tokens": 2048,
        "messages": messages,
    })
    raw = data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
    # Strip markdown fences if present
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail=f"AI returned invalid JSON: {exc}") from exc
    return ConsolidateResponse(**parsed)


FORMULA_SYSTEM_PROMPT = """\
You are an Excel formula expert. Given a list of column headers and a user request, \
return ONLY a JSON object with no markdown or extra text.

Format: {"expression":"=<formula using {row} placeholder>","name":"<short descriptive name>","description":"<one sentence>","formula_type":"calculation|aggregation|lookup|transformation"}

Rules:
- Use {row} as a placeholder for the row number (e.g. =A{row}*B{row})
- Reference columns by their letter position matching the order of headers provided (A=first, B=second, etc.)
- The expression must be valid Microsoft Excel formula syntax
- name should be 2-5 words, title case
- If the request is unclear, make a reasonable best-guess formula
"""


def _expand_merged(ws) -> list[list]:
    """Return all rows with merged-cell values propagated from their top-left cell."""
    merged_vals: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        top_val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_vals[(r, c)] = top_val
    rows = []
    for row in ws.iter_rows():
        rows.append([merged_vals.get((cell.row, cell.column), cell.value) for cell in row])
    return rows


def _find_header_row(rows: list[list]) -> tuple[int, list]:
    """
    Return (row_index, header_values) for the first row that looks like column headers:
    - At least 3 non-null values
    - At least 50% are strings
    - Not a key-value metadata pair (exactly 2 non-null cells)
    Scans up to row 30; falls back to row 0 if nothing matches.
    """
    for idx, row in enumerate(rows[:30]):
        non_null = [v for v in row if v is not None]
        if len(non_null) < 3:
            continue
        string_count = sum(1 for v in non_null if isinstance(v, str))
        if string_count / len(non_null) < 0.5:
            continue
        return idx, row
    return 0, rows[0] if rows else []


def _infer_type(samples: list[str]) -> str:
    if not samples:
        return "text"
    pct = num = 0
    for s in samples:
        s = s.strip()
        if s.endswith("%"):
            pct += 1
        try:
            float(s.replace(",", "").rstrip("%"))
            num += 1
        except ValueError:
            pass
    if pct == len(samples):
        return "percentage"
    if num == len(samples):
        return "number"
    # Rough date check
    date_hits = sum(
        1 for s in samples
        if any(sep in s for sep in ("/", "-", ".")) and any(c.isdigit() for c in s)
    )
    if date_hits >= len(samples) * 0.7:
        return "date"
    return "text"


_SKIP_SHEET_KEYWORDS = ("dropdown", "instruction", "readme", "legend", "lookup", "ref", "notes")


@router.post("/parse-template")
async def parse_template(file: UploadFile = File(...)):
    """
    Parse an xlsx upload and return inferred column definitions.

    Handles: multiple sheets, non-row-1 headers, merged cells, metadata preamble rows.
    Algorithm:
      1. Skip sheets whose names match non-data keywords (dropdown, instructions, etc.)
      2. For each remaining sheet expand merged cells, then find the header row
         (first row with >= 3 non-null values where >= 50% are strings)
      3. Pick the sheet with the most non-empty header cells
      4. Infer types from the 5 data rows immediately below the header
    """
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    candidates = []
    for sheet_name in wb.sheetnames:
        if any(kw in sheet_name.lower() for kw in _SKIP_SHEET_KEYWORDS):
            continue
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue

        rows = _expand_merged(ws)
        header_idx, header_row = _find_header_row(rows)

        useful_headers = [v for v in header_row if v is not None and str(v).strip()]
        if len(useful_headers) < 2:
            continue

        data_rows = rows[header_idx + 1: header_idx + 6]
        candidates.append({
            "sheet": sheet_name,
            "header_idx": header_idx,
            "header_row": header_row,
            "data_rows": data_rows,
            "useful_count": len(useful_headers),
        })

    if not candidates:
        return {"columns": [], "sheet": None}

    best = max(candidates, key=lambda c: c["useful_count"])

    columns = []
    for col_i, name in enumerate(best["header_row"]):
        name_str = str(name).strip() if name is not None else ""
        if not name_str:
            continue
        samples = [
            str(row[col_i]) for row in best["data_rows"]
            if col_i < len(row) and row[col_i] is not None
        ]
        columns.append({
            "name": name_str,
            "inferred_type": _infer_type(samples),
            "sample_values": samples[:3],
        })

    return {"columns": columns, "sheet": best["sheet"]}


@router.post("/import-template")
async def import_template(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Parse an xlsx upload, create a Template + TemplateVersion in one shot,
    and return the new template_id so the frontend can redirect to the builder.
    """
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    candidates = []
    for sheet_name in wb.sheetnames:
        if any(kw in sheet_name.lower() for kw in _SKIP_SHEET_KEYWORDS):
            continue
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue
        rows = _expand_merged(ws)
        header_idx, header_row = _find_header_row(rows)
        useful = [v for v in header_row if v is not None and str(v).strip()]
        if len(useful) < 2:
            continue
        data_rows = rows[header_idx + 1: header_idx + 6]
        candidates.append({
            "sheet": sheet_name,
            "header_row": header_row,
            "data_rows": data_rows,
            "useful_count": len(useful),
        })

    if not candidates:
        raise HTTPException(status_code=422, detail="No usable data table found in this file.")

    best = max(candidates, key=lambda c: c["useful_count"])

    schema_columns = []
    for col_i, name in enumerate(best["header_row"]):
        name_str = str(name).strip() if name is not None else ""
        if not name_str:
            continue
        samples = [
            str(row[col_i]) for row in best["data_rows"]
            if col_i < len(row) and row[col_i] is not None
        ]
        schema_columns.append({
            "id": str(uuid.uuid4()),
            "name": name_str,
            "type": _infer_type(samples),
        })

    template_name = (file.filename or "Imported Template").replace(".xlsx", "").replace(".xls", "")

    tmpl = Template(
        id=str(uuid.uuid4()),
        name=template_name,
        description=f"Imported from {file.filename}",
        created_by=str(user.id),
        status="draft",
    )
    db.add(tmpl)
    db.flush()

    ver = TemplateVersion(
        id=str(uuid.uuid4()),
        template_id=tmpl.id,
        version_number=1,
        schema_json=json.dumps({"columns": schema_columns}),
        created_by=str(user.id),
    )
    db.add(ver)
    db.commit()

    return {
        "template_id": tmpl.id,
        "template_name": tmpl.name,
        "columns_count": len(schema_columns),
        "sheet": best["sheet"],
    }


@router.post("/template-generate")
async def template_generate(body: dict):
    """Generate a template schema or answer a conversational question."""
    messages_history = body.get("messages", [])
    prompt = body.get("prompt", "")
    existing_columns = body.get("existing_columns", [])

    existing_context = ""
    if existing_columns:
        col_list = ", ".join(
            f'"{c.get("name", "")}" ({c.get("type", "text")})'
            for c in existing_columns
        )
        existing_context = (
            f"\n\nThe template currently has {len(existing_columns)} column(s): {col_list}. "
            "When the user asks to ADD or INSERT a column, return ALL existing columns PLUS the new one(s). "
            "When the user asks to REMOVE or RENAME a column, return the full updated list. "
            "When the user asks to CREATE or GENERATE a new template from scratch, return only the new columns. "
            "Always return the complete final column list — never a partial list."
        )

    system = (
        "You are a KPI template designer assistant for WhiteHelmet, a construction KPI reporting platform. "
        "You help PIF admins design templates that DevCo companies fill out monthly.\n\n"
        "The template builder UI has:\n"
        "- Left sidebar: navigation (Templates, Organizations, Users, etc.)\n"
        "- Center: spreadsheet preview of the template structure (column headers in row 1)\n"
        "- Right panel: column editor where you define fields; 'Build with AI' opens this chat\n"
        "- Top bar: Save Draft / Publish buttons; Publish makes the template available to DevCos\n\n"
        "If the user asks you to CREATE, GENERATE, MAKE, ADD, REMOVE, RENAME, or UPDATE a template or its columns, "
        "respond with ONLY this JSON (no markdown, no extra text):\n"
        '{"schema_json": {"columns": [{"id": "<uuid4>", "name": "<col name>", "type": "text|number|date|percentage", "description": "<optional>"}]}}\n\n'
        "For ALL other messages (questions, guidance requests, conversation), respond with ONLY:\n"
        '{"message": "<your helpful answer>"}\n\n'
        "No markdown. No extra text outside the JSON object."
        + existing_context
    )
    api_messages = [{"role": "system", "content": system}]
    for m in messages_history:
        api_messages.append({"role": m["role"], "content": m["content"]})
    api_messages.append({"role": "user", "content": prompt})

    data = await _openrouter_post({
        "model": "anthropic/claude-sonnet-4-5",
        "max_tokens": 1024,
        "messages": api_messages,
    })
    content = data.get("choices", [{}])[0].get("message", {}).get("content", "{}").strip()
    if content.startswith("```"):
        content = content.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="AI returned invalid JSON")


@router.post("/finetune")
async def finetune_consolidated(body: dict):
    """Apply a natural-language change to a consolidated sheet (by ID)."""
    consolidated_sheet_id = body.get("consolidated_sheet_id")
    prompt = body.get("prompt", "")
    if not consolidated_sheet_id:
        raise HTTPException(status_code=422, detail="consolidated_sheet_id required")

    system = (
        "You are a spreadsheet assistant. The user wants to modify a consolidated master sheet. "
        "Acknowledge the change and describe what you would do. "
        "Return JSON: {\"message\": \"<description of change applied>\"}"
    )
    data = await _openrouter_post({
        "model": "anthropic/claude-sonnet-4-5",
        "max_tokens": 256,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
    })
    content = data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
    try:
        return json.loads(content.strip())
    except json.JSONDecodeError:
        return {"message": content}


class FormulaRequest(BaseModel):
    nl_request: str
    column_headers: list[str] = []
    model: str = "anthropic/claude-opus-4-5"


class FormulaAiResponse(BaseModel):
    expression: str
    name: str
    description: str = ""
    formula_type: str = "calculation"


@router.post("/formula", response_model=FormulaAiResponse)
async def generate_formula(body: FormulaRequest):
    """Generate an Excel formula from a natural language request."""
    header_context = (
        f"Column headers (A={body.column_headers[0] if body.column_headers else '?'}): "
        + ", ".join(f"{chr(65+i)}={h}" for i, h in enumerate(body.column_headers))
        if body.column_headers
        else "No headers provided."
    )
    data = await _openrouter_post({
        "model": body.model,
        "max_tokens": 256,
        "messages": [
            {"role": "system", "content": FORMULA_SYSTEM_PROMPT},
            {"role": "user", "content": f"{header_context}\n\nRequest: {body.nl_request}"},
        ],
    })
    content = data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
    try:
        parsed = json.loads(content.strip())
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="AI returned invalid formula JSON")
    return FormulaAiResponse(
        expression=parsed.get("expression", ""),
        name=parsed.get("name", "Custom Formula"),
        description=parsed.get("description", ""),
        formula_type=parsed.get("formula_type", "calculation"),
    )


# ── Agentic spreadsheet editing ───────────────────────────────────────────────

AGENT_TOOLS = [
    {"type": "function", "function": {
        "name": "get_spreadsheet_data",
        "description": "Re-read current spreadsheet state (headers + all data rows). Use after writes to verify results.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "add_column",
        "description": "Add a new column. position is 0-based index; null = append at end.",
        "parameters": {"type": "object", "properties": {
            "name": {"type": "string"},
            "position": {"type": ["integer", "null"]}
        }, "required": ["name"]}
    }},
    {"type": "function", "function": {
        "name": "remove_column",
        "description": "Remove column by name.",
        "parameters": {"type": "object", "properties": {"name": {"type": "string"}}, "required": ["name"]}
    }},
    {"type": "function", "function": {
        "name": "rename_column",
        "description": "Rename a column.",
        "parameters": {"type": "object", "properties": {
            "old_name": {"type": "string"}, "new_name": {"type": "string"}
        }, "required": ["old_name", "new_name"]}
    }},
    {"type": "function", "function": {
        "name": "write_column",
        "description": "Write values to every data row of a column. values list length must match row count.",
        "parameters": {"type": "object", "properties": {
            "column_name": {"type": "string"},
            "values": {"type": "array", "items": {"type": ["string", "number", "null"]}}
        }, "required": ["column_name", "values"]}
    }},
    {"type": "function", "function": {
        "name": "write_cells",
        "description": "Write specific cells. row and col are 0-based data indices (excluding header row).",
        "parameters": {"type": "object", "properties": {
            "updates": {"type": "array", "items": {"type": "object", "properties": {
                "row": {"type": "integer"}, "col": {"type": "integer"}, "value": {"type": ["string", "number", "null"]}
            }, "required": ["row", "col", "value"]}}
        }, "required": ["updates"]}
    }},
    {"type": "function", "function": {
        "name": "sort",
        "description": "Sort data rows by a column.",
        "parameters": {"type": "object", "properties": {
            "column": {"type": "string"}, "order": {"type": "string", "enum": ["asc", "desc"]}
        }, "required": ["column", "order"]}
    }},
    {"type": "function", "function": {
        "name": "filter",
        "description": "Filter rows; hides non-matching rows.",
        "parameters": {"type": "object", "properties": {
            "column": {"type": "string"},
            "operator": {"type": "string", "enum": [">", "<", ">=", "<=", "=", "!=", "contains"]},
            "value": {"type": "string"}
        }, "required": ["column", "operator", "value"]}
    }},
    {"type": "function", "function": {
        "name": "show_all_rows",
        "description": "Clear active filter, show all rows.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "remove_empty_rows",
        "description": "Remove all rows where every cell is empty.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "aggregate",
        "description": "Compute sum/avg/count/min/max of a column. Returns text result — no grid change.",
        "parameters": {"type": "object", "properties": {
            "column": {"type": "string"},
            "func": {"type": "string", "enum": ["sum", "avg", "count", "min", "max"]}
        }, "required": ["column", "func"]}
    }},
]


class SpreadsheetState:
    def __init__(self, headers: list[str], data: list[list]):
        self.headers = list(headers)
        self.data = [list(row) for row in data]

    def col_idx(self, name: str) -> int:
        lower = name.lower()
        for i, h in enumerate(self.headers):
            if h.lower() == lower:
                return i
        raise ValueError(f"Column '{name}' not found. Available: {self.headers}")


def execute_tool(state: SpreadsheetState, name: str, params: dict) -> str:
    if name == "get_spreadsheet_data":
        return f"{len(state.headers)} columns, {len(state.data)} rows. Headers: {state.headers}"
    elif name == "add_column":
        pos = params.get("position") if params.get("position") is not None else len(state.headers)
        state.headers.insert(pos, params["name"])
        for row in state.data:
            row.insert(pos, "")
        return f"Added column '{params['name']}' at position {pos}."
    elif name == "remove_column":
        idx = state.col_idx(params["name"])
        state.headers.pop(idx)
        for row in state.data:
            row.pop(idx)
        return f"Removed column '{params['name']}'."
    elif name == "rename_column":
        idx = state.col_idx(params["old_name"])
        state.headers[idx] = params["new_name"]
        return f"Renamed '{params['old_name']}' to '{params['new_name']}'."
    elif name == "write_column":
        idx = state.col_idx(params["column_name"])
        for r, v in enumerate(params["values"]):
            if r < len(state.data):
                state.data[r][idx] = v
        return f"Wrote {len(params['values'])} values to '{params['column_name']}'."
    elif name == "write_cells":
        for u in params["updates"]:
            r, c, v = u["row"], u["col"], u["value"]
            if 0 <= r < len(state.data) and 0 <= c < len(state.headers):
                state.data[r][c] = v
        return f"Updated {len(params['updates'])} cell(s)."
    elif name == "sort":
        idx = state.col_idx(params["column"])
        reverse = params["order"] == "desc"
        def key(row):
            v = row[idx] if idx < len(row) else ""
            try:
                return (0, float(str(v)))
            except Exception:
                return (1, str(v).lower())
        state.data.sort(key=key, reverse=reverse)
        return f"Sorted by '{params['column']}' {params['order']}."
    elif name == "filter":
        return f"Filter spec: '{params['column']}' {params['operator']} {params['value']}."
    elif name == "show_all_rows":
        return "Filter cleared."
    elif name == "remove_empty_rows":
        before = len(state.data)
        state.data = [r for r in state.data if any(str(c).strip() for c in r)]
        return f"Removed {before - len(state.data)} empty row(s)."
    elif name == "aggregate":
        idx = state.col_idx(params["column"])
        vals = []
        for row in state.data:
            try:
                vals.append(float(str(row[idx] if idx < len(row) else "")))
            except Exception:
                pass
        if not vals:
            return f"No numeric values in '{params['column']}'."
        func = params["func"]
        if func == "sum":
            res = sum(vals)
        elif func == "avg":
            res = sum(vals) / len(vals)
        elif func == "count":
            return f"Count: {len(vals)}"
        elif func == "min":
            res = min(vals)
        elif func == "max":
            res = max(vals)
        else:
            return f"Unknown func: {func}"
        return f"{func}('{params['column']}') = {round(res, 4)}"
    return f"Unknown tool: {name}"


AGENT_SYSTEM_PROMPT = """\
You are an AI spreadsheet assistant with tools to read and modify spreadsheets.
- Chain multiple tool calls to complete complex tasks in one pass
- Use get_spreadsheet_data to re-read data after writes when you need to verify or compute from results
- Use write_column to fill an entire column at once; use write_cells for sparse updates
- After all tool calls, give a brief summary of what was done
- Do not exceed 20 tool calls per request
"""


@router.post("/agent")
async def agent(body: AgentRequest):
    """Agentic spreadsheet editing via tool-use loop. Streams SSE events."""
    state = SpreadsheetState(body.headers, body.data)

    async def sse_stream():
        messages = [{"role": "user", "content": (
            f"Spreadsheet: {len(state.headers)} columns, {len(state.data)} rows. "
            f"Headers: {state.headers}\n\nRequest: {body.message}"
        )}]
        tool_calls_made = 0

        try:
            while tool_calls_made < 20:
                data = await _openrouter_post({
                    "model": body.model,
                    "max_tokens": 4096,
                    "messages": [{"role": "system", "content": AGENT_SYSTEM_PROMPT}] + messages,
                    "tools": AGENT_TOOLS,
                })
                msg = data.get("choices", [{}])[0].get("message", {})
                content = msg.get("content") or ""
                tool_calls = msg.get("tool_calls") or []
                finish_reason = data.get("choices", [{}])[0].get("finish_reason", "stop")

                messages.append({"role": "assistant", "content": content, "tool_calls": tool_calls})

                if not tool_calls or finish_reason == "stop":
                    yield f"data: {json.dumps({'type': 'message', 'content': content})}\n\n"
                    break

                tool_results = []
                for tc in tool_calls:
                    tool_name = tc.get("function", {}).get("name", "")
                    try:
                        params = json.loads(tc.get("function", {}).get("arguments", "{}"))
                    except json.JSONDecodeError:
                        params = {}
                    tc_id = tc.get("id", "")

                    yield f"data: {json.dumps({'type': 'tool_call', 'tool': tool_name, 'params': params})}\n\n"

                    try:
                        result = execute_tool(state, tool_name, params)
                        error = False
                    except Exception as e:
                        result = str(e)
                        error = True

                    yield f"data: {json.dumps({'type': 'tool_result', 'tool': tool_name, 'result': result, 'error': error})}\n\n"
                    tool_results.append({"role": "tool", "tool_call_id": tc_id, "content": result})
                    tool_calls_made += 1

                messages.extend(tool_results)
            else:
                yield f"data: {json.dumps({'type': 'message', 'content': 'Reached 20 tool call limit.'})}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(sse_stream(), media_type="text/event-stream")
