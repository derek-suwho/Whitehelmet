"""Template routes — CRUD, versioning, status management."""

from __future__ import annotations

import json
import shutil
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse as FastAPIFileResponse
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.dependencies import get_current_user, verify_csrf
from app.core.paths import resolve_path, to_relative
from app.db.session import get_db
from app.models.consolidated_sheet import ConsolidatedSheet
from app.models.profile import Profile
from app.models.template import Template
from app.models.template_version import TemplateVersion
from app.schemas.templates import (
    ConsolidatedSheetResponse,
    TemplateCreate,
    TemplateResponse,
    TemplateVersionCreate,
    TemplateVersionResponse,
)

router = APIRouter(
    prefix="/api/templates",
    tags=["templates"],
    dependencies=[Depends(get_current_user)],
)


@router.get("", response_model=list[TemplateResponse])
def list_templates(
    type: str | None = Query(None, alias="type"),
    project_id: str | None = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Template)
    if type:
        q = q.filter(Template.template_type == type)
    if project_id:
        q = q.filter(Template.project_id == project_id)
    return q.order_by(Template.updated_at.desc()).all()


@router.get("/master", response_model=list[TemplateResponse])
def list_master_templates(db: Session = Depends(get_db)):
    """Return all master templates, newest first."""
    return db.query(Template).filter(Template.template_type == "master").order_by(Template.updated_at.desc()).all()


@router.post(
    "", response_model=TemplateResponse, status_code=status.HTTP_201_CREATED, dependencies=[Depends(verify_csrf)]
)
def create_template(body: TemplateCreate, user: Profile = Depends(get_current_user), db: Session = Depends(get_db)):
    tmpl = Template(
        id=str(uuid.uuid4()),
        name=body.name,
        description=body.description,
        created_by=str(user.id),
        status="draft",
        template_type=body.template_type,
        project_id=body.project_id,
    )
    db.add(tmpl)
    db.commit()
    db.refresh(tmpl)
    return tmpl


@router.get("/{template_id}", response_model=TemplateResponse)
def get_template(template_id: str, db: Session = Depends(get_db)):
    tmpl = db.query(Template).filter(Template.id == template_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    return tmpl


@router.patch("/{template_id}", response_model=TemplateResponse, dependencies=[Depends(verify_csrf)])
def update_template(template_id: str, body: dict, db: Session = Depends(get_db)):
    tmpl = db.query(Template).filter(Template.id == template_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    if "name" in body and body["name"]:
        tmpl.name = body["name"]
    if "description" in body:
        tmpl.description = body["description"]
    if "template_type" in body and body["template_type"] in ("subcontractor", "master"):
        tmpl.template_type = body["template_type"]
    db.commit()
    db.refresh(tmpl)
    return tmpl


@router.delete("/{template_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(verify_csrf)])
def delete_template(template_id: str, db: Session = Depends(get_db)):
    from app.models.template_version import TemplateVersion
    from app.models.template_assignment import TemplateAssignment
    from app.models.template_formula import TemplateFormula
    from app.models.consolidated_sheet import ConsolidatedSheet

    tmpl = db.query(Template).filter(Template.id == template_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")

    version_ids = [row[0] for row in db.query(TemplateVersion.id).filter(TemplateVersion.template_id == template_id).all()]
    if version_ids:
        db.query(TemplateAssignment).filter(TemplateAssignment.template_version_id.in_(version_ids)).delete(synchronize_session=False)
        db.query(TemplateFormula).filter(TemplateFormula.template_version_id.in_(version_ids)).delete(synchronize_session=False)
        db.query(TemplateVersion).filter(TemplateVersion.template_id == template_id).delete(synchronize_session=False)
    db.query(ConsolidatedSheet).filter(ConsolidatedSheet.template_id == template_id).delete(synchronize_session=False)
    db.delete(tmpl)
    db.commit()


@router.patch("/{template_id}/status", response_model=TemplateResponse, dependencies=[Depends(verify_csrf)])
def update_status(template_id: str, body: dict, db: Session = Depends(get_db)):
    tmpl = db.query(Template).filter(Template.id == template_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    new_status = body.get("status")
    if new_status not in ("draft", "active", "deprecated"):
        raise HTTPException(status_code=422, detail="Invalid status")
    tmpl.status = new_status
    db.commit()
    db.refresh(tmpl)
    return tmpl


@router.get("/{template_id}/download-xlsx")
def download_template_xlsx(template_id: str, db: Session = Depends(get_db)):
    """Serve the latest uploaded xlsx file for a template (admin editor use)."""
    ver = (
        db.query(TemplateVersion)
        .filter(TemplateVersion.template_id == template_id)
        .order_by(TemplateVersion.version_number.desc())
        .first()
    )
    if not ver or not ver.file_path:
        raise HTTPException(status_code=404, detail="No xlsx file for this template")
    path = resolve_path(ver.file_path)
    if not path or not path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FastAPIFileResponse(
        path=str(path),
        filename=f"template_{template_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/{template_id}/versions", response_model=list[TemplateVersionResponse])
def list_versions(template_id: str, db: Session = Depends(get_db)):
    return (
        db.query(TemplateVersion)
        .filter(TemplateVersion.template_id == template_id)
        .order_by(TemplateVersion.version_number.desc())
        .all()
    )


@router.post(
    "/{template_id}/versions",
    response_model=TemplateVersionResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(verify_csrf)],
)
def save_version(
    template_id: str,
    body: TemplateVersionCreate,
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tmpl = db.query(Template).filter(Template.id == template_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")

    last = (
        db.query(TemplateVersion)
        .filter(TemplateVersion.template_id == template_id)
        .order_by(TemplateVersion.version_number.desc())
        .first()
    )
    next_version = (last.version_number if last else 0) + 1

    ver = TemplateVersion(
        id=str(uuid.uuid4()),
        template_id=template_id,
        version_number=next_version,
        schema_json=json.dumps(body.schema_data),
        created_by=str(user.id),
    )
    db.add(ver)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Version conflict — please retry") from None
    db.refresh(ver)
    return ver


@router.post(
    "/upload-xlsx",
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(verify_csrf)],
)
async def upload_xlsx_template(
    name: str = Query(..., description="Template name"),
    template_type: str = Query("subcontractor", description="Template type: subcontractor or master"),
    project_id: str | None = Query(None),
    file: UploadFile = File(...),
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create a new template + version directly from an uploaded xlsx file."""
    from app.core.config import get_settings
    settings = get_settings()

    raw = await file.read()
    upload_dir = Path(settings.upload_dir) / "templates"
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_id = str(uuid.uuid4())
    dest = upload_dir / f"{file_id}.xlsx"

    if template_type == "master":
        # Master templates: save raw bytes to preserve exact file structure
        dest.write_bytes(raw)
    else:
        # Subcontractor templates: strip utility sheets (Dropdown, Sheet1, Quality, etc.)
        import openpyxl, io as _io
        wb_in = openpyxl.load_workbook(_io.BytesIO(raw))
        _UTILITY_SHEETS = {"dropdown", "sheet1", "quality", "cover", "instructions"}
        sheets_to_keep = [s for s in wb_in.sheetnames if s.lower() not in _UTILITY_SHEETS]
        if sheets_to_keep:
            for sheet_name in wb_in.sheetnames[:]:
                if sheet_name.lower() in _UTILITY_SHEETS:
                    del wb_in[sheet_name]
        wb_in.save(str(dest))

    # Create Template record
    tmpl = Template(
        id=str(uuid.uuid4()),
        name=name,
        template_type=template_type if template_type in ("subcontractor", "master") else "subcontractor",
        status="active",
        created_by=str(user.id),
        project_id=project_id,
    )
    db.add(tmpl)
    db.flush()

    # Create TemplateVersion with empty schema + file path
    ver = TemplateVersion(
        id=str(uuid.uuid4()),
        template_id=tmpl.id,
        version_number=1,
        schema_json='{"columns":[]}',
        file_path=to_relative(dest),
        created_by=str(user.id),
    )
    db.add(ver)
    db.commit()
    db.refresh(tmpl)
    db.refresh(ver)

    return {"template_id": tmpl.id, "version_id": ver.id, "name": tmpl.name}


@router.get("/consolidations/{sheet_id}/download")
def download_consolidated(sheet_id: str, db: Session = Depends(get_db)):
    sheet = db.query(ConsolidatedSheet).filter(ConsolidatedSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Consolidated sheet not found")
    path = resolve_path(sheet.file_path)
    if not path or not path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FastAPIFileResponse(
        path=str(path),
        filename=f"consolidated_{sheet_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/{template_id}/consolidations", response_model=list[ConsolidatedSheetResponse])
def list_consolidations(template_id: str, db: Session = Depends(get_db)):
    return (
        db.query(ConsolidatedSheet)
        .filter(ConsolidatedSheet.template_id == template_id)
        .order_by(ConsolidatedSheet.generated_at.desc())
        .all()
    )
