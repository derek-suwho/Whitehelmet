"""Admin routes — user management and consolidation progress."""

from __future__ import annotations

import json
import logging
import re
import uuid as _uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

from app.core.config import get_settings
from app.core.dependencies import get_current_user, verify_csrf
from app.core.paths import resolve_path, to_relative
from app.core.rbac import require_admin
from app.db.session import get_db
from app.models.consolidated_sheet import ConsolidatedSheet
from app.models.profile import Profile
from app.models.project import Project
from app.models.project_member import ProjectMember
from app.models.submission import Submission
from app.models.template import Template
from app.models.template_assignment import TemplateAssignment
from app.models.template_formula import TemplateFormula
from app.models.template_version import TemplateVersion
from app.schemas.admin import UpdateOrgRequest, UpdateRoleRequest, UserWithOrgResponse
from app.schemas.subcontractor import ConsolidationProgressResponse, OrgSubmissionStatus
from app.schemas.template_formula import (
    TemplateFormulaCreate,
    TemplateFormulaResponse,
    TemplateFormulaUpdate,
)

router = APIRouter(
    prefix="/api/admin",
    tags=["admin"],
    dependencies=[Depends(get_current_user)],
)


@router.get("/users", response_model=list[UserWithOrgResponse], dependencies=[Depends(require_admin)])
def list_users(db: Session = Depends(get_db)):
    profiles = db.query(Profile).order_by(Profile.display_name).all()
    from app.models.project_member import ProjectMember
    from app.models.project import Project
    result = []
    all_memberships = db.query(ProjectMember).all()
    all_projects = {proj.id: proj for proj in db.query(Project).all()}
    memberships_by_user: dict = {}
    for pm in all_memberships:
        memberships_by_user.setdefault(str(pm.user_id), []).append(pm)

    for p in profiles:
        user_memberships = memberships_by_user.get(str(p.id), [])
        projects = []
        for pm in user_memberships:
            proj = all_projects.get(pm.project_id)
            if proj:
                projects.append({
                    "project_id": proj.id,
                    "project_name": proj.name,
                    "membership_id": pm.id,
                    "participant_role": pm.participant_role,
                })
        project_name = projects[0]["project_name"] if projects else None
        participant_role = projects[0]["participant_role"] if projects else None
        result.append(UserWithOrgResponse(
            id=str(p.id),
            display_name=p.display_name,
            email=p.email,
            role=p.role,
            org_id=str(p.org_id) if p.org_id else None,
            org_name=p.org_name,
            project_name=project_name,
            participant_role=participant_role,
            projects=projects,
        ))
    return result


@router.patch(
    "/orgs/{org_id}/name",
    dependencies=[Depends(require_admin), Depends(verify_csrf)],
)
def update_org_name(org_id: str, body: UpdateOrgRequest, db: Session = Depends(get_db)):
    """Set org_name for all profiles with this org_id."""
    updated = db.query(Profile).filter(Profile.org_id == org_id).update(
        {"org_name": body.org_name.strip()}, synchronize_session=False
    )
    db.commit()
    return {"ok": True, "updated_users": updated}


@router.patch(
    "/users/{user_id}/role",
    response_model=UserWithOrgResponse,
    dependencies=[Depends(require_admin), Depends(verify_csrf)],
)
def update_user_role(user_id: str, body: UpdateRoleRequest, db: Session = Depends(get_db)):
    from app.models.project_member import ProjectMember
    user = db.query(Profile).filter(Profile.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.role = body.role
    if body.participant_role and body.role == "participant":
        valid_roles = {"focal", "member", "viewer"}
        if body.participant_role in valid_roles:
            db.query(ProjectMember).filter(ProjectMember.user_id == user_id).update(
                {"participant_role": body.participant_role}, synchronize_session=False
            )
    db.commit()
    db.refresh(user)
    # Return with projects populated
    memberships = db.query(ProjectMember).filter(ProjectMember.user_id == user_id).all()
    from app.models.project import Project
    all_projects = {p.id: p for p in db.query(Project).all()}
    projects = []
    for pm in memberships:
        proj = all_projects.get(pm.project_id)
        if proj:
            projects.append({"project_id": proj.id, "project_name": proj.name,
                             "membership_id": pm.id, "participant_role": pm.participant_role})
    participant_role = projects[0]["participant_role"] if projects else None
    return UserWithOrgResponse(
        id=str(user.id), display_name=user.display_name, email=user.email,
        role=user.role, org_id=str(user.org_id) if user.org_id else None,
        project_name=projects[0]["project_name"] if projects else None,
        participant_role=participant_role, projects=projects,
    )


@router.get("/orgs", dependencies=[Depends(require_admin)])
def list_orgs(db: Session = Depends(get_db)):
    """Return distinct orgs derived from participant profiles, with member names."""
    participants = db.query(Profile).filter(
        Profile.role == "participant", Profile.org_id.isnot(None)
    ).order_by(Profile.display_name).all()
    orgs: dict = {}
    for p in participants:
        oid = str(p.org_id)
        if oid not in orgs:
            orgs[oid] = {"org_id": oid, "org_name": p.org_name, "members": []}
        orgs[oid]["members"].append({"display_name": p.display_name, "email": p.email})
    return list(orgs.values())


class AssignToOrgRequest(BaseModel):
    template_version_id: str
    org_ids: list[str]
    deadline: str | None = None


@router.post("/assign-template-to-org", status_code=201, dependencies=[Depends(require_admin), Depends(verify_csrf)])
def assign_template_to_org(
    body: AssignToOrgRequest,
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Assign a template version to one or more orgs (org_id, not project_id)."""
    from datetime import datetime
    deadline_dt = datetime.fromisoformat(body.deadline) if body.deadline else None
    created = []
    for org_id in body.org_ids:
        existing = db.query(TemplateAssignment).filter(
            TemplateAssignment.org_id == org_id,
            TemplateAssignment.template_version_id == body.template_version_id,
            TemplateAssignment.assigned_to_user_id.is_(None),
        ).first()
        if existing:
            continue
        a = TemplateAssignment(
            id=str(_uuid.uuid4()),
            template_version_id=body.template_version_id,
            org_id=org_id,
            assigned_by=str(user.id),
            deadline=deadline_dt,
            submission_type="template",
            status="pending",
            assigned_to_user_id=None,
        )
        db.add(a)
        created.append(org_id)
    db.commit()
    return {"ok": True, "assigned_to": created}


@router.get(
    "/templates/{template_id}/consolidation-progress",
    response_model=ConsolidationProgressResponse,
    dependencies=[Depends(require_admin)],
)
def get_consolidation_progress(
    template_id: str,
    project_id: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """Return per-org submission status.

    When project_id is supplied (master-template context) the response lists
    each project member and their most recent submission.  Otherwise the
    existing assignment-based logic is used.
    """
    if project_id:
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        members = db.query(ProjectMember).filter(ProjectMember.project_id == project_id).all()

        _proj_member_ids = [m.user_id for m in members]
        _proj_profiles = (
            {p.id: p for p in db.query(Profile).filter(Profile.id.in_(_proj_member_ids)).all()}
            if _proj_member_ids
            else {}
        )

        # Pre-fetch all assignments keyed by org_id for fast per-member lookup
        from sqlalchemy import or_ as _or_
        _all_org_ids = list({str(p.org_id) for p in _proj_profiles.values() if p.org_id})
        _asgn_filter = _or_(
            TemplateAssignment.org_id == project_id,
            *(TemplateAssignment.org_id == oid for oid in _all_org_ids),
        ) if _all_org_ids else (TemplateAssignment.org_id == project_id)
        _all_assignments = db.query(TemplateAssignment).filter(_asgn_filter).all()
        # Map org_id -> list of assignment IDs
        _org_assignment_ids: dict[str, list[str]] = {}
        for a in _all_assignments:
            _org_assignment_ids.setdefault(a.org_id, []).append(a.id)

        org_rows: list[OrgSubmissionStatus] = []
        submitted_count = 0
        member_count = 0

        for m in members:
            user = _proj_profiles.get(m.user_id)
            if not user:
                continue
            member_count += 1

            # Only look at assignments for this member's org (+ project-level)
            member_asgn_ids = _org_assignment_ids.get(project_id, [])
            if user.org_id:
                member_asgn_ids = list(set(member_asgn_ids + _org_assignment_ids.get(str(user.org_id), [])))

            sub = None
            if member_asgn_ids:
                sub = (
                    db.query(Submission)
                    .filter(
                        Submission.assignment_id.in_(member_asgn_ids),
                        Submission.submitted_by == str(user.id),
                    )
                    .order_by(Submission.submitted_at.desc())
                    .first()
                )

            if sub:
                submitted_count += 1
                org_rows.append(
                    OrgSubmissionStatus(
                        org_id=str(user.id),
                        org_name=user.display_name,
                        assignment_id=sub.assignment_id,
                        assignment_status="submitted",
                        submission_id=sub.id,
                        submitted_at=sub.submitted_at,
                        file_name=sub.file_name,
                    )
                )
            else:
                org_rows.append(
                    OrgSubmissionStatus(
                        org_id=str(user.id),
                        org_name=user.display_name,
                        assignment_id=None,
                        assignment_status="pending",
                        submission_id=None,
                        submitted_at=None,
                        file_name=None,
                    )
                )

        return ConsolidationProgressResponse(
            template_id=template_id,
            template_version_id=None,
            total_orgs=member_count,
            submitted_count=submitted_count,
            all_submitted=submitted_count == member_count and member_count > 0,
            orgs=org_rows,
        )

    # Assignment-based mode (subcontractor templates)
    latest_ver = (
        db.query(TemplateVersion)
        .filter(TemplateVersion.template_id == template_id)
        .order_by(TemplateVersion.version_number.desc())
        .first()
    )

    assignments = []
    if latest_ver:
        assignments = (
            db.query(TemplateAssignment)
            .filter(
                TemplateAssignment.template_version_id == latest_ver.id,
                TemplateAssignment.submission_type == "template",
            )
            .all()
        )

    org_rows = []
    submitted_count = 0

    for a in assignments:
        project = db.query(Project).filter(Project.id == a.org_id).first()
        subs = (
            db.query(Submission).filter(Submission.assignment_id == a.id).order_by(Submission.submitted_at.desc()).all()
        )

        if subs:
            # One row per submitter so every file is visible and selectable
            for sub in subs:
                submitted_count += 1
                submitter = db.query(Profile).filter(Profile.id == sub.submitted_by).first()
                display = submitter.display_name if submitter else (project.name if project else a.org_id)
                org_rows.append(
                    OrgSubmissionStatus(
                        org_id=sub.submitted_by or a.org_id,
                        org_name=display,
                        assignment_id=a.id,
                        assignment_status="submitted",
                        submission_id=sub.id,
                        submitted_at=sub.submitted_at,
                        file_name=sub.file_name,
                    )
                )
        else:
            org_rows.append(
                OrgSubmissionStatus(
                    org_id=a.org_id,
                    org_name=project.name if project else a.org_id,
                    assignment_id=a.id,
                    assignment_status=a.status,
                    submission_id=None,
                    submitted_at=None,
                    file_name=None,
                )
            )

    return ConsolidationProgressResponse(
        template_id=template_id,
        template_version_id=latest_ver.id if latest_ver else None,
        total_orgs=len(org_rows),
        submitted_count=submitted_count,
        all_submitted=submitted_count == len(assignments) and len(assignments) > 0,
        orgs=org_rows,
    )


@router.get(
    "/projects/{project_id}/submission-overview",
    dependencies=[Depends(require_admin)],
)
def get_project_submission_overview(project_id: str, db: Session = Depends(get_db)):
    """Return cross-template submission counts for a project.

    Overall: total expected (members × templates) and total received.
    Per-template: submitted_count and total_members for each template assigned to the project.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    members = db.query(ProjectMember).filter(ProjectMember.project_id == project_id).all()
    total_members = len(members)
    member_ids = [str(m.user_id) for m in members]

    # All template version assignments for this project (project-wide org_id assignments)
    project_assignments = db.query(TemplateAssignment).filter(TemplateAssignment.org_id == project_id).all()

    # Resolve unique templates via template_version → template
    seen_template_ids: set[str] = set()
    templates_progress = []

    for assignment in project_assignments:
        if not assignment.template_version_id:
            continue
        version = db.query(TemplateVersion).filter(TemplateVersion.id == assignment.template_version_id).first()
        if not version:
            continue
        template = db.query(Template).filter(Template.id == version.template_id).first()
        if not template or template.id in seen_template_ids:
            continue
        seen_template_ids.add(template.id)

        # All assignment IDs for this template in this project
        t_assignment_ids = [
            a.id
            for a in db.query(TemplateAssignment)
            .filter(
                TemplateAssignment.org_id == project_id,
                TemplateAssignment.template_version_id == version.id,
            )
            .all()
        ]

        submitted_count = sum(
            1 for uid in member_ids
            if db.query(Submission).filter(
                Submission.assignment_id.in_(t_assignment_ids),
                Submission.submitted_by == uid,
            ).first()
        )

        templates_progress.append(
            {
                "template_id": template.id,
                "template_name": template.name,
                "total_members": total_members,
                "submitted_count": submitted_count,
                "all_submitted": submitted_count == total_members and total_members > 0,
            }
        )

    total_expected = total_members * len(templates_progress)
    total_submitted = sum(t["submitted_count"] for t in templates_progress)

    return {
        "total_members": total_members,
        "total_expected": total_expected,
        "total_submitted": total_submitted,
        "all_submitted": total_submitted == total_expected and total_expected > 0,
        "templates": templates_progress,
    }




@router.post(
    "/templates/{template_id}/consolidate-submissions",
    dependencies=[Depends(require_admin), Depends(verify_csrf)],
)
async def consolidate_submissions(
    template_id: str,
    body: dict,
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Load project submission files, run AI schema unification, save ConsolidatedSheet."""
    from app.api.routes.ai import _ai_post

    settings = get_settings()
    project_id: str | None = body.get("project_id")
    submission_ids: list | None = body.get("submission_ids")
    report_name: str | None = body.get("name")
    report_period: str | None = body.get("period")

    if submission_ids:
        submissions = db.query(Submission).filter(Submission.id.in_(submission_ids)).all()
    elif project_id:
        from sqlalchemy import or_ as _or_
        # Include both legacy project-scoped (org_id=project_id) and new org-based assignments
        _member_ids = [m.user_id for m in db.query(ProjectMember).filter(ProjectMember.project_id == project_id).all()]
        _member_org_ids = list({str(p.org_id) for p in db.query(Profile).filter(Profile.id.in_(_member_ids)).all() if p.org_id})
        _org_filter = _or_(
            TemplateAssignment.org_id == project_id,
            *(TemplateAssignment.org_id == oid for oid in _member_org_ids),
        ) if _member_org_ids else (TemplateAssignment.org_id == project_id)
        aids = [a.id for a in db.query(TemplateAssignment).filter(_org_filter).all()]
        submissions = db.query(Submission).filter(Submission.assignment_id.in_(aids)).all() if aids else []
    else:
        raise HTTPException(status_code=400, detail="project_id or submission_ids required")

    if not submissions:
        raise HTTPException(status_code=400, detail="No submissions to consolidate")

    from app.api.routes.ai import _expand_merged

    # Sheets to always skip (utility / formula / metadata sheets)
    _SKIP_SHEETS = (
        "dropdown",
        "instruction",
        "readme",
        "legend",
        "lookup",
        "ref",
        "notes",
        "calculation sheet",
        "calculation",
        "formula",
        "sheet1",
        "company infomation",
        "company information",
        "company info",
        "wbs",
        "contracts",
        "construction cost",
        "milestones",
        "progress",
        "assets",
        "observations",
    )
    # Preferred DevCo input sheet names — checked in order; first match wins
    _PREFERRED = ("ard", "quality", "data", "input", "kpi", "hse", "safety", "metrics")

    def _best_header_row(rows: list) -> tuple:
        """Return the most likely column-name row in the first 30 rows.

        Scoring: most DISTINCT non-empty strings wins. Ties broken by shortest
        average string length — column names (e.g. '#', 'Level 1') are short,
        description/preamble rows use full sentences, so the column-names row
        wins even when both rows have the same number of distinct strings."""
        best_idx, best_row, best_count, best_avg = 0, rows[0] if rows else [], 0, float("inf")
        for idx, row in enumerate(rows[:30]):
            strings = [str(v).strip() for v in row if v is not None and isinstance(v, str) and str(v).strip()]
            distinct = len(set(strings))
            avg_len = sum(len(s) for s in strings) / len(strings) if strings else float("inf")
            if distinct > best_count or (distinct == best_count and avg_len < best_avg):
                best_count, best_idx, best_row, best_avg = distinct, idx, row, avg_len
        return best_idx, best_row

    def _pick_sheet(wb):
        """Return (sheet_name, rows, header_idx, header_row) for the DevCo input sheet."""
        candidates = {}
        for name in wb.sheetnames:
            if any(kw in name.lower() for kw in _SKIP_SHEETS):
                continue
            ws = wb[name]
            if ws.max_row is None or ws.max_row < 2:
                continue
            rows = _expand_merged(ws)
            hidx, hrow = _best_header_row(rows)
            useful = [v for v in hrow if v is not None and str(v).strip()]
            if len(useful) < 3:
                continue
            data = [r for r in rows[hidx + 1 :] if any(c is not None and str(c).strip() for c in r)]
            if not data:
                continue
            candidates[name] = (name, rows, hidx, hrow, len(useful))

        # Prefer known input sheet names (order matters: "ard" before "quality")
        for keyword in _PREFERRED:
            for name, val in candidates.items():
                if keyword in name.lower():
                    return val
        # Fall back to the sheet with the most header columns
        if candidates:
            return max(candidates.values(), key=lambda c: c[4])
        return None

    # Parse each xlsx
    file_schemas = []
    all_file_data = []
    failed_files: list[str] = []
    for sub in submissions:
        path = resolve_path(sub.processed_file_path) or resolve_path(sub.file_path)
        if not path or not path.exists():
            import logging as _logging

            _logging.getLogger(__name__).warning(
                "consolidate_submissions: file missing for submission %s: %s", sub.id, path
            )
            failed_files.append(sub.file_name)
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            picked = _pick_sheet(wb)
            if not picked:
                import logging as _logging

                _logging.getLogger(__name__).warning(
                    "consolidate_submissions: no usable sheet in submission %s (%s)", sub.id, sub.file_name
                )
                failed_files.append(sub.file_name)
                continue
            _, rows, header_idx, header_row, _ = picked

            # Truncate verbose header strings — KPI templates use full sentences as column names
            headers = [str(h).strip()[:80] if h is not None else "" for h in header_row]

            def _nonempty(v) -> bool:
                return v is not None and str(v).strip() != ""

            def _coerce(v):
                """Preserve native numeric types; only stringify text."""
                if v is None:
                    return None
                if isinstance(v, (int, float)):
                    return v
                s = str(v).strip()
                if s == "":
                    return None
                try:
                    return int(s)
                except ValueError:
                    pass
                try:
                    return float(s)
                except ValueError:
                    return s

            # Keep rows with ≥3 non-empty cells AND >1 distinct value
            # (excludes sparse sequence-number rows and label rows like "KPI Ratings" repeated)
            data_rows = [
                [_coerce(c) for c in r]
                for r in rows[header_idx + 1 :]
                if (sum(1 for c in r if _nonempty(c)) >= 3 and len({str(c) for c in r if _nonempty(c)}) > 1)
            ]
            if not data_rows:
                continue

            file_schemas.append({"name": sub.file_name, "headers": headers, "sample_rows": data_rows[:5]})
            all_file_data.append({"name": sub.file_name, "headers": headers, "all_rows": data_rows})
        except Exception as exc:
            import logging as _logging

            _logging.getLogger(__name__).warning(
                "consolidate_submissions: failed to parse %s (sub %s): %s", sub.file_name, sub.id, exc
            )
            failed_files.append(sub.file_name)
            continue

    if not all_file_data:
        raise HTTPException(status_code=400, detail="Could not read any submission files")

    # ── Version-consistency guard (AR-3, AR-9) ───────────────────────
    # Only enforced when formulas exist — legacy/formula-free consolidation
    # proceeds without requiring version metadata on every submission.
    from sqlalchemy import distinct

    submission_version_ids = (
        db.query(distinct(TemplateAssignment.template_version_id))
        .join(Submission, Submission.assignment_id == TemplateAssignment.id)
        .filter(Submission.id.in_([s.id for s in submissions]))
        .all()
    )
    non_null_version_ids = [vid for (vid,) in submission_version_ids if vid]
    has_null_versions = any(vid is None for (vid,) in submission_version_ids)

    # Resolve template version and formulas (best-effort — may be None for legacy data)
    template_version = (
        db.query(TemplateVersion)
        .filter(TemplateVersion.id == non_null_version_ids[0])
        .first()
    ) if non_null_version_ids else None

    t_formulas = []
    if template_version:
        t_formulas = (
            db.query(TemplateFormula)
            .filter(TemplateFormula.template_version_id == template_version.id)
            .all()
        )

    # Guards only fire when formulas exist — no-formula consolidation works with any data
    if t_formulas:
        if has_null_versions:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Some submissions have no template version assigned. "
                    "Formula-aware consolidation requires every submission to be "
                    "linked to a specific template version. Please backfill the "
                    "template_version_id on all assignments before consolidating."
                ),
            )

        if len(non_null_version_ids) > 1:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Cannot consolidate submissions from different template versions. "
                    f"Found versions: {non_null_version_ids}. "
                    "Please ensure all submissions use the same template version, "
                    "or reassign them before consolidating."
                ),
            )

    # AI schema unification
    system_prompt = (
        "You are a spreadsheet schema unifier. "
        "Given headers and sample rows from multiple Excel files, "
        "return ONLY a valid JSON object — no markdown, no extra text:\n"
        '{"unified_headers":["Source File","<col>",...],'
        '"mappings":[{"file":"<name>","column_map":{"<src>":"<unified>"}}]}\n\n'
        "Rules:\n"
        '- Always include "Source File" as the first unified header.\n'
        "- Merge columns that represent the same concept under one standard name.\n"
        "- Include every column that appears in at least one file.\n"
        "- column_map must cover every source column in that file."
    )
    ai_data = await _ai_post(
        {
            "model": "anthropic/claude-sonnet-4-5",
            "max_tokens": 8192,
            "messages": [
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": json.dumps(
                        [
                            {"name": f["name"], "headers": f["headers"], "sample_rows": f["sample_rows"][:3]}
                            for f in file_schemas
                        ]
                    ),
                },
            ],
        }
    )

    raw_ai = ai_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
    raw_ai = raw_ai.strip()
    if raw_ai.startswith("```"):
        raw_ai = raw_ai.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
    try:
        schema = json.loads(raw_ai)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail=f"AI returned invalid schema: {exc}") from exc

    unified_headers = schema.get("unified_headers", ["Source File"])
    mappings = {m["file"]: m.get("column_map", {}) for m in schema.get("mappings", [])}

    # Merge rows — emit None for missing values so Excel leaves the cell blank
    merged_rows = []
    for f in all_file_data:
        col_map = mappings.get(f["name"], {})
        unified_to_idx: dict[str, int] = {}
        for src, unified in col_map.items():
            try:
                unified_to_idx[unified] = f["headers"].index(src)
            except ValueError:
                pass
        for row in f["all_rows"]:
            out = [f["name"]]
            for h in unified_headers[1:]:
                idx = unified_to_idx.get(h)
                val = row[idx] if idx is not None and idx < len(row) else None
                out.append(val)
            merged_rows.append(out)

    # Write output xlsx — values are already native Python types (int/float/str/None)
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Consolidated"
    ws_out.append(unified_headers)
    for row in merged_rows:
        ws_out.append(row)

    out_dir = Path(settings.upload_dir) / "consolidated"
    out_dir.mkdir(parents=True, exist_ok=True)
    sheet_id = str(_uuid.uuid4())
    out_path = out_dir / f"{sheet_id}.xlsx"
    wb_out.save(out_path)

    # ── Apply template formulas as Excel formula strings ──────────────
    if t_formulas:
        from app.services.formula_executor import FormulaExecutor

        t_hr = template_version.header_row or 1
        file_path = template_version.file_path

        if not file_path:
            raise HTTPException(
                status_code=500,
                detail=(
                    f"Template version {template_version.id} has no xlsx snapshot. "
                    "Cannot remap formulas without a workbook artifact. "
                    "Re-save the template version to generate one."
                ),
            )

        # Build template header maps per sheet
        template_h2c_by_sheet: dict[str, dict[str, int]] = {}
        active_sheet_name: str | None = None

        t_path = resolve_path(file_path)
        if t_path and t_path.exists():
            t_wb = openpyxl.load_workbook(t_path, data_only=True)
            active_sheet_name = t_wb.active.title
            for ws_name in t_wb.sheetnames:
                t_ws = t_wb[ws_name]
                h2c: dict[str, int] = {}
                for ci in range(1, t_ws.max_column + 1):
                    hdr = t_ws.cell(row=t_hr, column=ci).value
                    if hdr:
                        h2c[str(hdr)] = ci
                template_h2c_by_sheet[ws_name] = h2c

        template_h2c = template_h2c_by_sheet.get(active_sheet_name or "", {})

        # Build consolidated header→col map (row 1)
        consol_h2c: dict[str, int] = {}
        for ci, h in enumerate(unified_headers, start=1):
            if h:
                consol_h2c[h] = ci

        # Build header_rename_map from AI's column_map
        header_rename_map: dict[str, str] = {}
        for file_map in mappings.values():
            for src, unified in file_map.items():
                if src != unified:
                    header_rename_map[src] = unified
            break  # one file's map is sufficient

        # Remap and write formulas
        wb_consol = openpyxl.load_workbook(out_path)
        ws_consol = wb_consol.active
        data_start = 2
        data_end = ws_consol.max_row
        single_cell_offset = 0

        # Filter formulas — reject non-active-sheet targeting
        skipped_sheets: set[str] = set()
        applicable_formulas = []
        for fm in t_formulas:
            fm_sheet = getattr(fm, "target_sheet", None)
            if fm_sheet and fm_sheet != active_sheet_name:
                skipped_sheets.add(fm_sheet)
                continue
            applicable_formulas.append(fm)

        if skipped_sheets:
            import logging
            logging.getLogger(__name__).warning(
                "Consolidation skipped %d formula(s) targeting non-active sheets: %s",
                len(t_formulas) - len(applicable_formulas),
                skipped_sheets,
            )

        # Pre-process: build a map of header names → column letter for replacing
        # literal header names that someone typed directly in formula expressions.
        def _normalize_expr_headers(expr: str, h2c: dict[str, int]) -> str:
            """Replace literal header names in expression with column letter + {row}."""
            from app.services.formula_executor import _col_index_to_letter
            for name in sorted(h2c.keys(), key=len, reverse=True):
                if name in expr:
                    letter = _col_index_to_letter(h2c[name])
                    expr = expr.replace(name, f"{letter}{{row}}")
            return expr

        for fm in applicable_formulas:
            fm_sheet = getattr(fm, "target_sheet", None)
            fm_h2c = template_h2c_by_sheet.get(fm_sheet, template_h2c) if fm_sheet else template_h2c

            remapped_target = FormulaExecutor.remap_target_column(
                fm.target_column, fm_h2c, consol_h2c, header_rename_map
            )

            target_col_idx = consol_h2c.get(remapped_target)
            if not target_col_idx:
                target_col_idx = ws_consol.max_column + 1
                ws_consol.cell(row=1, column=target_col_idx, value=remapped_target)
                consol_h2c[remapped_target] = target_col_idx

            # Pre-process: resolve header names typed directly in expression
            expr = _normalize_expr_headers(fm.expression, fm_h2c)

            remapped_expr = FormulaExecutor.remap_expression(
                expr, fm_h2c, consol_h2c, header_rename_map
            )

            if fm.formula_type == "column":
                for row in range(data_start, data_end + 1):
                    cell_formula = remapped_expr.replace("{row}", str(row))
                    ws_consol.cell(row=row, column=target_col_idx, value=cell_formula)
            else:
                single_cell_offset += 1
                write_row = data_end + 1 + single_cell_offset
                cell_formula = remapped_expr.replace("{row}", str(write_row))
                ws_consol.cell(row=write_row, column=target_col_idx, value=cell_formula)

        wb_consol.save(out_path)

    auto_name = report_name or f"Consolidation – {datetime.utcnow().strftime('%b %d, %Y')}"
    sheet = ConsolidatedSheet(
        id=sheet_id,
        template_id=template_id,
        project_id=project_id,
        name=auto_name,
        period=report_period,
        file_path=to_relative(out_path),
        generated_by=str(user.id),
    )
    db.add(sheet)
    db.commit()

    return {
        "consolidated_sheet_id": sheet_id,
        "file_path": str(out_path),
        "template_count": len(all_file_data),
        "freeform_count": 0,
        "name": auto_name,
        "period": report_period,
        "failed_files": failed_files,
    }


# ── Assignment Lock / Unlock ─────────────────────────────────────────────────


class AssignmentLockResponse(BaseModel):
    id: str
    status: str
    locked_at: datetime | None = None
    locked_by: str | None = None

    model_config = {"from_attributes": True}


# ── Master report management ──────────────────────────────────────────────────


class ConsolidatedSheetResponse(BaseModel):
    id: str
    template_id: str
    project_id: str | None
    name: str | None
    period: str | None
    generated_by: str | None
    generated_at: datetime

    model_config = {"from_attributes": True}


@router.delete("/assignments/{assignment_id}", status_code=204, dependencies=[Depends(require_admin), Depends(verify_csrf)])
def delete_assignment(assignment_id: str, db: Session = Depends(get_db)):
    """Remove a template assignment (and its submissions)."""
    a = db.query(TemplateAssignment).filter(TemplateAssignment.id == assignment_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.query(Submission).filter(Submission.assignment_id == assignment_id).delete(synchronize_session=False)
    db.delete(a)
    db.commit()


@router.post("/assignments/{assignment_id}/lock", response_model=AssignmentLockResponse)
def lock_assignment(
    assignment_id: str,
    user: Profile = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Lock a submission — no further uploads accepted from the DevCo."""
    a = db.query(TemplateAssignment).filter(TemplateAssignment.id == assignment_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if a.status == "locked":
        raise HTTPException(status_code=409, detail="Assignment already locked")
    a.status = "locked"
    a.locked_at = datetime.now(timezone.utc)
    a.locked_by = str(user.id)
    db.commit()
    db.refresh(a)
    return a


@router.post("/assignments/{assignment_id}/unlock", response_model=AssignmentLockResponse)
def unlock_assignment(
    assignment_id: str,
    user: Profile = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Unlock a submission — allows DevCo to resubmit."""
    a = db.query(TemplateAssignment).filter(TemplateAssignment.id == assignment_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    a.status = "submitted"
    a.locked_at = None
    a.locked_by = None
    db.commit()
    db.refresh(a)
    return a


# ── Template stamp helper ─────────────────────────────────────────────────────

_STAMP_PROJECT_KW = ("project name", "project names", "project")
_STAMP_SOURCE_KW = ("source file", "source", "devco", "file name")
_RATE_SUFFIXES = (" RATE", " FREQUENCY RATE", " FREQUENCY", " INDEX", " RATIO")
_TEXT_CRITERIA_RE = re.compile(r"[Ss]core\s+(?:\d+|[Zz]ero)", re.IGNORECASE)


def _parse_scoring_criteria_text(text: str) -> list[dict]:
    """Parse free-text scoring criteria into rule dicts for _apply_scoring.

    Handles formats such as:
      "1. 0 : Score 100%"          → exact match 0 → score 100
      "> 0 and ≤ 0.01: Score 90%"  → (0, 0.01] → score 90
      "0.01 and ≤ 0.05: Score 80"  → [0.01, 0.05] → score 80
      "> 0.1: Score Zero"           → > 0.1 → score 0
      "< 1.0: Score 70%"            → < 1.0 → score 70
      "3.0 - 5.0: Score 90%"        → [3.0, 5.0] → score 90
    Works for both newline-separated and inline numbered-list formats.
    """
    rules: list[dict] = []

    for raw_cond, raw_score in re.findall(
        r"([^\n:]+?)\s*:\s*[Ss]core\s+(0*\d{1,3}(?:\.\d+)?|[Zz]ero)%?(?!\d)",
        text,
    ):
        cond = re.sub(r"^\s*\d+[.)]\s+", "", raw_cond).strip()
        if not cond:
            continue

        raw_score = raw_score.strip()
        score = 0 if raw_score.lower() == "zero" else int(float(raw_score.lstrip("0") or "0"))
        rule: dict = {"score": score}

        nums = [float(n) for n in re.findall(r"\d+(?:\.\d+)?", cond)]

        if ("-" in cond or "\u2013" in cond) and len(nums) >= 2:
            # "X - Y" inclusive range
            rule["min"] = nums[0]
            rule["max"] = nums[1]
        elif "and" in cond.lower() and len(nums) >= 2:
            # "> X and ≤ Y" / "X and ≤ Y"
            rule["min"] = nums[0]
            rule["max"] = nums[1]
        elif re.match(r"^\s*[>≥]", cond):
            # "> X" or ">= X"
            if nums:
                rule["min"] = nums[0]
        elif re.match(r"^\s*[<≤]", cond):
            # "< X" or "<= X"
            if nums:
                rule["max"] = nums[0]
        elif len(nums) == 1:
            # exact "= X" or bare "X"
            rule["min"] = nums[0]
            rule["max"] = nums[0]
        else:
            continue

        rules.append(rule)

    return rules


def _stamp_with_template(
    consol_xlsx: bytes,
    template_path,
    header_row: int,
    template_h2c: dict[str, int],
    scoring_by_tmpl_col: "dict[int, list] | None" = None,
) -> bytes:
    """Reshape a consolidated xlsx to match the master template structure.

    Only formula-computed (rate) columns are written into the template —
    raw input columns (manhours, incident counts, etc.) are excluded.
    Project name and source file are appended as extras after the template columns.
    Formula strings in the consolidated workbook are evaluated via xlcalculator
    before copying so they don't break in the template context.
    """
    import io as _io

    wb_final = openpyxl.load_workbook(_io.BytesIO(consol_xlsx))
    ws_final = wb_final.active

    # ── Identify formula (rate) columns before evaluation ────────────────────
    formula_cols: set[int] = set()
    for ci in range(1, ws_final.max_column + 1):
        for ri in range(2, min(ws_final.max_row + 1, 5)):  # sample first few rows
            v = ws_final.cell(row=ri, column=ci).value
            if isinstance(v, str) and v.startswith("="):
                formula_cols.add(ci)
                break

    # ── Pre-fill None with 0 in non-formula numeric columns ──────────────────
    # xlcalculator treats blank cells as errors (not 0 like Excel), so formulas
    # like =D2*200000/E2 where D2 is blank return errors instead of 0.
    for _ci in range(1, ws_final.max_column + 1):
        if _ci in formula_cols:
            continue
        _is_numeric = any(
            isinstance(ws_final.cell(row=_ri, column=_ci).value, (int, float))
            for _ri in range(2, ws_final.max_row + 1)
        )
        if _is_numeric:
            for _ri in range(2, ws_final.max_row + 1):
                if ws_final.cell(row=_ri, column=_ci).value is None:
                    ws_final.cell(row=_ri, column=_ci).value = 0

    # ── Evaluate formula strings via xlcalculator ────────────────────────────
    try:
        from xlcalculator import ModelCompiler, Evaluator as _Eval
        from app.services.formula_executor import _col_index_to_letter as _cil
        _buf = _io.BytesIO()
        wb_final.save(_buf)
        _buf.seek(0)
        _compiler = ModelCompiler()
        _model = _compiler.read_and_parse_archive(_buf, ignore_sheets=[])
        _evaluator = _Eval(_model)
        _sheet = ws_final.title
        for ci in formula_cols:
            for ri in range(2, ws_final.max_row + 1):
                cell = ws_final.cell(row=ri, column=ci)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    addr = f"{_sheet}!{_cil(ci)}{ri}"
                    try:
                        computed = _evaluator.evaluate(addr)
                        # xlcalculator's Number(0.0) == None, so use `is not None` (identity)
                        cell.value = float(computed) if computed is not None else None
                    except Exception:
                        cell.value = None
    except Exception:
        pass  # proceed with raw values if xlcalculator unavailable

    # ── Build header map ─────────────────────────────────────────────────────
    final_h2c: dict[str, int] = {}
    for ci in range(1, ws_final.max_column + 1):
        h = ws_final.cell(row=1, column=ci).value
        if h:
            final_h2c[str(h)] = ci

    t_wb_stamp = openpyxl.load_workbook(template_path)
    t_ws_stamp = t_wb_stamp.active

    tmpl_upper_to_col: dict[str, int] = {k.upper(): v for k, v in template_h2c.items()}

    # ── Map columns to template columns ──────────────────────────────────────
    # Non-formula cols: exact header match only.
    # Formula (rate) cols: exact match first, then fuzzy rate-suffix match.
    consol_to_tmpl: dict[int, int] = {}
    for h, ci in final_h2c.items():
        tmpl_ci = tmpl_upper_to_col.get(h.upper())
        if tmpl_ci is None and ci in formula_cols:
            # Fuzzy match: strip trailing parenthetical abbreviations like (LTIFR),
            # then strip rate suffixes, then find template col whose header contains the base.
            # e.g. "Lost Time Injury Frequency Rate (LTIFR)" → "Lost Time Injury Frequency Rate"
            #      → base "LOST TIME INJURY" matches "Lost Time Injury incidents"
            import re as _re
            h_upper = _re.sub(r"\s*\([^)]*\)\s*$", "", h.upper()).strip()
            for suffix in _RATE_SUFFIXES:
                if h_upper.endswith(suffix):
                    base = h_upper[: -len(suffix)].strip()
                    if not base:
                        continue
                    best_ci: int | None = None
                    best_extra: float = float("inf")
                    for tmpl_h, tcol in tmpl_upper_to_col.items():
                        if base in tmpl_h:
                            extra = len(tmpl_h) - len(base)
                            if extra < best_extra:
                                best_extra, best_ci = extra, tcol
                    if best_ci is not None:
                        tmpl_ci = best_ci
                        break
        if tmpl_ci is not None:
            consol_to_tmpl[ci] = tmpl_ci

    # ── Extras: project name and source file only ────────────────────────────
    unmapped: list[tuple[str, int]] = []
    for h, ci in final_h2c.items():
        if ci in consol_to_tmpl:
            continue
        hl = h.lower()
        if any(kw in hl for kw in _STAMP_PROJECT_KW) or any(kw in hl for kw in _STAMP_SOURCE_KW):
            unmapped.append((h, ci))

    def _extra_sort_key(item: tuple[str, int]) -> int:
        hl = item[0].lower()
        if any(kw in hl for kw in _STAMP_PROJECT_KW):
            return 0
        return 1

    unmapped.sort(key=_extra_sort_key)

    # ── Safe-write helper (skips read-only MergedCell objects) ───────────────
    from openpyxl.cell.cell import MergedCell as _MergedCell

    def _safe_write(ws, row, col, value, reset_format=False):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, _MergedCell):
            return
        cell.value = value
        if reset_format and value is not None:
            cell.number_format = "General"

    used_tmpl_cols = set(template_h2c.values())
    extra_col_map: dict[int, int] = {}

    # Find template column designated for project names in any non-header row
    _pn_designated_col: int | None = None
    for _r in range(1, header_row):
        for _c in range(1, t_ws_stamp.max_column + 1):
            _v = t_ws_stamp.cell(row=_r, column=_c).value
            if isinstance(_v, str) and any(kw in _v.lower() for kw in _STAMP_PROJECT_KW):
                if _c not in used_tmpl_cols:
                    _pn_designated_col = _c
                break
        if _pn_designated_col is not None:
            break

    # Project names → template-designated col if found, else leftmost free col
    pn_col: int
    if _pn_designated_col is not None:
        pn_col = _pn_designated_col
    else:
        pn_col = 1
        while pn_col in used_tmpl_cols:
            pn_col += 1

    # Source file → col 1 if free and before pn_col, else after last template col
    if 1 not in used_tmpl_cols and pn_col != 1:
        src_col = 1
    else:
        src_col = (max(template_h2c.values()) + 1) if template_h2c else (pn_col + 1)
        while src_col in used_tmpl_cols or src_col == pn_col:
            src_col += 1

    right_col = (max(template_h2c.values()) + 1) if template_h2c else (max(pn_col, src_col) + 1)
    while right_col in used_tmpl_cols or right_col in {pn_col, src_col}:
        right_col += 1

    for h, ci in unmapped:
        hl = h.lower()
        if any(kw in hl for kw in _STAMP_PROJECT_KW):
            _safe_write(t_ws_stamp, header_row, pn_col, h)
            extra_col_map[ci] = pn_col
        elif any(kw in hl for kw in _STAMP_SOURCE_KW):
            _safe_write(t_ws_stamp, header_row, src_col, h)
            extra_col_map[ci] = src_col
        else:
            _safe_write(t_ws_stamp, header_row, right_col, h)
            extra_col_map[ci] = right_col
            right_col += 1

    # ── Clear stale data rows in template ────────────────────────────────────
    for row in range(header_row + 1, t_ws_stamp.max_row + 1):
        for col in range(1, t_ws_stamp.max_column + 1):
            _safe_write(t_ws_stamp, row, col, None)

    # ── Write rate values + extras into template, unhiding each data row ─────
    _tmpl_scoring = scoring_by_tmpl_col or {}

    for row_offset, src_row in enumerate(range(2, ws_final.max_row + 1), start=1):
        tgt_row = header_row + row_offset
        t_ws_stamp.row_dimensions[tgt_row].hidden = False
        for consol_ci, tmpl_ci in consol_to_tmpl.items():
            val = ws_final.cell(row=src_row, column=consol_ci).value
            _safe_write(t_ws_stamp, tgt_row, tmpl_ci, val, reset_format=True)
            # Write score to adjacent column (tmpl_ci + 1) if rules exist for this template col
            if _tmpl_scoring and val is not None:
                from app.services.formula_executor import _apply_scoring
                _rules = _tmpl_scoring.get(tmpl_ci)
                if _rules:
                    try:
                        _score = _apply_scoring(float(val), _rules)
                        _safe_write(t_ws_stamp, tgt_row, tmpl_ci + 1, _score, reset_format=True)
                    except (TypeError, ValueError):
                        pass
        for consol_ci, tmpl_ci in extra_col_map.items():
            val = ws_final.cell(row=src_row, column=consol_ci).value
            _safe_write(t_ws_stamp, tgt_row, tmpl_ci, val)

    buf = _io.BytesIO()
    t_wb_stamp.save(buf)
    return buf.getvalue()


# ── Apply Formula Library to Consolidated Sheet ──────────────────────────────


def _col_idx_to_letter(index: int) -> str:
    """1-based column index → Excel column letter (1→A, 27→AA)."""
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result


def _decode_param_name(param: str) -> str:
    """Decode underscore-encoded param to expression form: __ → '. ', _ → ' '."""
    return param.replace("__", ". ").replace("_", " ")


def _normalize_for_match(s: str) -> str:
    """Normalize text for fuzzy matching: lowercase, collapse periods/spaces."""
    return re.sub(r"[\s.]+", " ", s.lower()).strip()


@router.post(
    "/consolidated-sheets/{sheet_id}/apply-formulas",
    dependencies=[Depends(require_admin), Depends(verify_csrf)],
)
def apply_library_formulas(
    sheet_id: str,
    body: dict,
    user: Profile = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Apply Formula Library formulas to a consolidated sheet.

    body: { "formulas": [
        { "formula_id": "...", "param_map": {"param_name": "Column Header"}, "output_name": "Rate Name" }
    ]}
    """
    from app.models.formula import Formula

    settings = get_settings()
    sheet = db.query(ConsolidatedSheet).filter(ConsolidatedSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Consolidated sheet not found")

    path = resolve_path(sheet.file_path)
    if not path or not path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    formula_configs = body.get("formulas", [])
    if not formula_configs:
        raise HTTPException(status_code=400, detail="No formulas provided")

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Build header→col_idx map (1-indexed)
    header_to_col: dict[str, int] = {}
    for ci in range(1, ws.max_column + 1):
        hdr = ws.cell(row=1, column=ci).value
        if hdr:
            header_to_col[str(hdr)] = ci

    data_start = 2
    data_end = ws.max_row
    applied = []

    for cfg in formula_configs:
        formula_id = cfg.get("formula_id")
        param_map = cfg.get("param_map", {})
        output_name = cfg.get("output_name", "")

        formula = db.query(Formula).filter(Formula.id == formula_id).first()
        if not formula:
            logger.warning("apply-formulas: formula_id=%s not found in DB", formula_id)
            continue

        # Parse parameters from JSON
        params = []
        if formula.parameters:
            try:
                params = json.loads(formula.parameters)
            except Exception:
                pass

        # Build param→column letter mapping
        param_to_letter: dict[str, str] = {}
        for param in params:
            col_name = param_map.get(param)
            if col_name and col_name in header_to_col:
                param_to_letter[param] = _col_idx_to_letter(header_to_col[col_name])

        logger.info("apply-formulas: %s | params=%s | param_map=%s | param_to_letter=%s | expression=%r",
                     formula.name, params, param_map, param_to_letter, formula.expression)

        if len(param_to_letter) != len(params):
            logger.warning("apply-formulas: SKIPPED %s — param_to_letter(%d) != params(%d)",
                           formula.name, len(param_to_letter), len(params))
            continue  # skip if not all params are mapped

        # Reuse existing output column or append new one
        out_name = output_name or formula.name
        if out_name in header_to_col:
            out_col_idx = header_to_col[out_name]
        else:
            out_col_idx = ws.max_column + 1
            ws.cell(row=1, column=out_col_idx, value=out_name)
            header_to_col[out_name] = out_col_idx

        # Build unified replacement map: expression term → column letter.
        # Params are underscore-encoded (__ = '. ', _ = ' ') so we decode
        # them to match the expression text. Also include sheet headers as
        # fallback for expressions that use current column names directly.
        replacements: dict[str, str] = {}
        for param in params:
            if param in param_to_letter:
                letter = param_to_letter[param]
                replacements[param] = letter
                decoded = _decode_param_name(param)
                if decoded != param:
                    replacements[decoded] = letter
        for hdr, col_idx in header_to_col.items():
            if hdr == out_name or hdr in replacements:
                continue
            replacements[hdr] = _col_idx_to_letter(col_idx)

        # Replace longest keys first to avoid partial matches
        sorted_keys = sorted(replacements, key=len, reverse=True)

        # Build normalized lookup for fallback matching (handles encoding
        # mismatches like "No. of N.M." vs decoded "No. of N M ")
        norm_replacements: dict[str, tuple[str, str]] = {}
        for key in sorted_keys:
            norm_key = _normalize_for_match(key)
            if norm_key not in norm_replacements:
                norm_replacements[norm_key] = (key, replacements[key])

        # Pre-compute: detect params whose names don't appear in the expression
        # (e.g. expression uses "No. of N.M." but param is "near_misses").
        # Build a fragment→letter map for last-resort substitution.
        used_params = set()
        for p in params:
            if p in formula.expression or _decode_param_name(p) in formula.expression:
                used_params.add(p)
        unused_mapped_letters = [
            param_to_letter[p]
            for p in params
            if p not in used_params and p in param_to_letter
        ]

        for row in range(data_start, data_end + 1):
            excel_expr = formula.expression
            for key in sorted_keys:
                if key in excel_expr:
                    excel_expr = excel_expr.replace(key, f"{replacements[key]}{row}")

            # Fallback 1: find unresolved text fragments and match via normalization
            remaining = re.findall(r"[A-Za-z][A-Za-z0-9. ]*[A-Za-z.]", excel_expr)
            for fragment in remaining:
                if re.fullmatch(r"[A-Z]+\d+", fragment):
                    continue
                norm_frag = _normalize_for_match(fragment)
                if norm_frag in norm_replacements:
                    _, letter = norm_replacements[norm_frag]
                    excel_expr = excel_expr.replace(fragment, f"{letter}{row}")

            # Fallback 2: assign unresolved fragments to unused param columns
            if unused_mapped_letters:
                unresolved = [
                    f for f in re.findall(r"[A-Za-z][A-Za-z0-9. ]*[A-Za-z.]", excel_expr)
                    if not re.fullmatch(r"[A-Z]+\d+", f)
                ]
                for i, fragment in enumerate(unresolved):
                    if i >= len(unused_mapped_letters):
                        break
                    excel_expr = excel_expr.replace(fragment, f"{unused_mapped_letters[i]}{row}")

            if not excel_expr.startswith("="):
                excel_expr = "=" + excel_expr
            ws.cell(row=row, column=out_col_idx, value=excel_expr)

        applied.append(output_name or formula.name)

    wb.save(path)
    return {"applied": applied, "count": len(applied)}


# ── Template Formula CRUD ────────────────────────────────────────────────────


@router.get(
    "/template-versions/{version_id}/formulas",
    response_model=list[TemplateFormulaResponse],
)
def list_formulas(
    version_id: str,
    user: Profile = Depends(require_admin),
    db: Session = Depends(get_db),
):
    return db.query(TemplateFormula).filter(TemplateFormula.template_version_id == version_id).all()


@router.post(
    "/template-versions/{version_id}/formulas",
    response_model=TemplateFormulaResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_formula(
    version_id: str,
    body: TemplateFormulaCreate,
    user: Profile = Depends(require_admin),
    db: Session = Depends(get_db),
):
    import uuid as _u

    f = TemplateFormula(
        id=str(_u.uuid4()),
        template_version_id=version_id,
        name=body.name,
        target_column=body.target_column,
        formula_type=body.formula_type,
        expression=body.expression,
        weight=body.weight,
        benchmark=body.benchmark,
        scoring_rules=json.dumps(body.scoring_rules) if body.scoring_rules is not None else None,
        created_by=str(user.id),
    )
    db.add(f)
    db.commit()
    db.refresh(f)
    return f


@router.put("/formulas/{formula_id}", response_model=TemplateFormulaResponse)
def update_formula(
    formula_id: str,
    body: TemplateFormulaUpdate,
    user: Profile = Depends(require_admin),
    db: Session = Depends(get_db),
):
    f = db.query(TemplateFormula).filter(TemplateFormula.id == formula_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Formula not found")
    for field, val in body.model_dump(exclude_unset=True).items():
        if field == "scoring_rules" and val is not None:
            val = json.dumps(val)
        setattr(f, field, val)
    db.commit()
    db.refresh(f)
    return f


@router.delete("/formulas/{formula_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_formula(
    formula_id: str,
    user: Profile = Depends(require_admin),
    db: Session = Depends(get_db),
):
    f = db.query(TemplateFormula).filter(TemplateFormula.id == formula_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Formula not found")
    db.delete(f)
    db.commit()


# ── Master report management ─────────────────────────────────────────────────


class RenameReportRequest(BaseModel):
    name: str


@router.get(
    "/projects/{project_id}/master-reports",
    response_model=list[ConsolidatedSheetResponse],
    dependencies=[Depends(require_admin)],
)
def list_master_reports(project_id: str, db: Session = Depends(get_db)):
    """List all consolidated master reports for a project, newest first."""
    return (
        db.query(ConsolidatedSheet)
        .filter(ConsolidatedSheet.project_id == project_id)
        .order_by(ConsolidatedSheet.generated_at.desc())
        .all()
    )


@router.patch(
    "/consolidated-sheets/{sheet_id}/rename",
    response_model=ConsolidatedSheetResponse,
    dependencies=[Depends(require_admin), Depends(verify_csrf)],
)
def rename_master_report(
    sheet_id: str,
    body: RenameReportRequest,
    db: Session = Depends(get_db),
):
    """Rename a consolidated master report."""
    sheet = db.query(ConsolidatedSheet).filter(ConsolidatedSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Report not found")
    sheet.name = body.name
    db.commit()
    db.refresh(sheet)
    return sheet


@router.get(
    "/consolidated-sheets/{sheet_id}/meta",
    dependencies=[Depends(require_admin)],
)
def get_consolidated_sheet_meta(sheet_id: str, db: Session = Depends(get_db)):
    sheet = db.query(ConsolidatedSheet).filter(ConsolidatedSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    return {"template_id": sheet.template_id, "project_id": sheet.project_id, "name": sheet.name}


@router.delete(
    "/consolidated-sheets/{sheet_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_admin), Depends(verify_csrf)],
)
def delete_master_report(sheet_id: str, db: Session = Depends(get_db)):
    """Delete a consolidated master report record and its file from disk."""
    sheet = db.query(ConsolidatedSheet).filter(ConsolidatedSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Report not found")
    try:
        _del_path = resolve_path(sheet.file_path)
        if _del_path:
            _del_path.unlink(missing_ok=True)
    except Exception:
        pass
    db.delete(sheet)
    db.commit()


@router.get(
    "/consolidated-sheets/{sheet_id}/stamped-template",
    dependencies=[Depends(require_admin)],
)
def download_stamped_template(sheet_id: str, db: Session = Depends(get_db)):
    """Return the consolidated sheet reshaped into its master template structure."""
    from fastapi.responses import Response as _Resp

    sheet = db.query(ConsolidatedSheet).filter(ConsolidatedSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Report not found")
    consol_path = resolve_path(sheet.file_path)
    if not consol_path or not consol_path.exists():
        raise HTTPException(status_code=404, detail="Consolidated file not found on disk")

    # Resolve the latest template version with an xlsx snapshot
    tv = (
        db.query(TemplateVersion)
        .filter(
            TemplateVersion.template_id == sheet.template_id,
            TemplateVersion.file_path.isnot(None),
        )
        .order_by(TemplateVersion.version_number.desc())
        .first()
    )
    if not tv:
        raise HTTPException(
            status_code=400,
            detail="No template version with an xlsx snapshot found. Re-save the template version first.",
        )

    t_path = resolve_path(tv.file_path)
    if not t_path or not t_path.exists():
        raise HTTPException(status_code=404, detail="Template xlsx file not found on disk")

    t_wb = openpyxl.load_workbook(t_path, data_only=True)
    t_ws = t_wb.active

    # Load consolidated headers to use as ground truth for finding the right
    # header row in the master template. The KPI template's header_row applies
    # to submissions; the master template xlsx may have a different structure.
    import io as _io2
    _consol_wb = openpyxl.load_workbook(_io2.BytesIO(consol_path.read_bytes()), data_only=True)
    _consol_ws = _consol_wb.active
    consol_header_upper: set[str] = set()
    for ci in range(1, _consol_ws.max_column + 1):
        v = _consol_ws.cell(row=1, column=ci).value
        if isinstance(v, str) and not v.startswith("#"):
            consol_header_upper.add(" ".join(v.split()).upper())

    # Pick the template row with the most matches against consolidated headers.
    # Normalize whitespace (collapse newlines/multiple spaces) so headers like
    # "Total Recordable \nIncident" match "Total Recordable Incident".
    template_h2c: dict[str, int] = {}
    t_hr = tv.header_row or 1
    best_overlap = -1
    for r in range(1, min(t_ws.max_row + 1, 20)):
        h2c: dict[str, int] = {}
        for ci in range(1, t_ws.max_column + 1):
            v = t_ws.cell(row=r, column=ci).value
            if isinstance(v, str) and not v.startswith("#") and len(v.strip()) > 2:
                h2c[" ".join(v.split())] = ci  # collapse all whitespace
        overlap = sum(1 for k in h2c if k.upper() in consol_header_upper)
        if overlap > best_overlap:
            best_overlap = overlap
            template_h2c = h2c
            t_hr = r

    if not template_h2c:
        raise HTTPException(status_code=400, detail="Template has no headers at the configured header row.")

    # Build scoring rules keyed by TEMPLATE COLUMN INDEX (reliable — avoids
    # mismatch between formula.target_column and AI-unified consolidated headers).
    # For each formula, match target_column against template headers (exact then
    # fuzzy rate-suffix), then store rules at that template column index.
    _scoring_by_tmpl_col: dict[int, list] = {}
    _tv_formulas = db.query(TemplateFormula).filter(TemplateFormula.template_version_id == tv.id).all()
    _tmpl_upper_map = {k.upper(): v for k, v in template_h2c.items()}
    import re as _re2
    for _f in _tv_formulas:
        if not _f.scoring_rules:
            continue
        try:
            _rules = json.loads(_f.scoring_rules)
        except (json.JSONDecodeError, TypeError):
            continue
        _tgt_upper = _f.target_column.upper()
        _tmpl_ci = _tmpl_upper_map.get(_tgt_upper)
        if _tmpl_ci is None:
            # Fuzzy: strip parenthetical abbrev then rate suffixes
            _h = _re2.sub(r"\s*\([^)]*\)\s*$", "", _tgt_upper).strip()
            for _sfx in (" RATE", " FREQUENCY RATE", " FREQUENCY", " INDEX", " RATIO"):
                if _h.endswith(_sfx):
                    _base = _h[: -len(_sfx)].strip()
                    _best_ci, _best_extra = None, float("inf")
                    for _th, _tci in _tmpl_upper_map.items():
                        if _base in _th:
                            _x = len(_th) - len(_base)
                            if _x < _best_extra:
                                _best_extra, _best_ci = _x, _tci
                    if _best_ci is not None:
                        _tmpl_ci = _best_ci
                        break
        if _tmpl_ci is not None:
            _scoring_by_tmpl_col[_tmpl_ci] = _rules

    # Supplement DB rules with scoring criteria parsed from template text
    # (rows above the header row).  DB rules take precedence if both exist.
    for _ci in range(1, t_ws.max_column + 1):
        if _ci in _scoring_by_tmpl_col:
            continue
        for _r in range(1, t_hr):
            _cell_val = t_ws.cell(row=_r, column=_ci).value
            if isinstance(_cell_val, str) and _TEXT_CRITERIA_RE.search(_cell_val):
                _parsed = _parse_scoring_criteria_text(_cell_val)
                if _parsed:
                    _scoring_by_tmpl_col[_ci] = _parsed
                    break

    stamped = _stamp_with_template(
        consol_path.read_bytes(), t_path, t_hr, template_h2c,
        scoring_by_tmpl_col=_scoring_by_tmpl_col or None,
    )
    raw_name = f"{sheet.name or sheet_id} - Template View.xlsx"
    filename = raw_name.encode("ascii", errors="replace").decode("ascii").replace("/", "-")
    return _Resp(
        content=stamped,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/consolidated-sheets/{sheet_id}/download",
    dependencies=[Depends(require_admin)],
)
def download_master_report(sheet_id: str, db: Session = Depends(get_db)):
    """Download the xlsx file for a consolidated master report."""
    sheet = db.query(ConsolidatedSheet).filter(ConsolidatedSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Report not found")
    path = resolve_path(sheet.file_path)
    if not path or not path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    filename = f"{sheet.name or sheet_id}.xlsx".replace("/", "-")
    return FileResponse(
        str(path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
