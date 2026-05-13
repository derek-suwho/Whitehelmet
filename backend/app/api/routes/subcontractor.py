"""Subcontractor portal routes — assignment list, template download, file upload."""

from __future__ import annotations

import json
import tempfile
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import FileResponse as FastAPIFileResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.dependencies import get_current_user, verify_csrf
from app.core.rbac import _is_admin, require_participant
from app.core.security import hash_file
from app.db.session import get_db
from app.models.profile import Profile
from app.models.project_member import ProjectMember
from app.models.submission import Submission
from app.models.template import Template
from app.models.template_assignment import TemplateAssignment
from app.models.template_formula import TemplateFormula
from app.models.template_version import TemplateVersion
from app.schemas.subcontractor import AssignmentForSubcontractor, SubmissionResponse
from app.services.formula_executor import FormulaExecutor

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}

router = APIRouter(
    prefix="/api/subcontractor",
    tags=["subcontractor"],
    dependencies=[Depends(require_participant)],
)


def _user_project_ids(user: Profile, db) -> list[str]:
    """Return all project IDs this user belongs to via ProjectMember."""
    memberships = db.query(ProjectMember).filter(ProjectMember.user_id == str(user.id)).all()
    return [m.project_id for m in memberships]


def _assignment_filter(user: Profile):
    """SQLAlchemy filter: assignment belongs to user's org (org-based) or is user-scoped (legacy)."""
    return or_(
        TemplateAssignment.org_id == str(user.org_id),
        TemplateAssignment.assigned_to_user_id == str(user.id),
    )


def _merge_into_template(submitted_bytes: bytes, template_version_id: str | None, db: Session) -> bytes | None:
    """Overlay submitted values onto the original template, preserving all cell styles.

    Returns merged bytes, or None if the merge can't be done (caller falls back to raw bytes).
    """
    if not template_version_id:
        return None
    try:
        import io
        import logging
        import openpyxl

        ver = db.query(TemplateVersion).filter(TemplateVersion.id == template_version_id).first()
        if not ver or not ver.file_path or not Path(ver.file_path).exists():
            return None

        template_wb = openpyxl.load_workbook(ver.file_path)
        submitted_wb = openpyxl.load_workbook(io.BytesIO(submitted_bytes), data_only=True)

        for sheet_name in submitted_wb.sheetnames:
            if sheet_name not in template_wb.sheetnames:
                continue
            t_ws = template_wb[sheet_name]
            s_ws = submitted_wb[sheet_name]
            for row in s_ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    t_cell = t_ws.cell(row=cell.row, column=cell.column)
                    # Never overwrite formula cells — keep template formulas intact
                    if isinstance(t_cell.value, str) and t_cell.value.startswith("="):
                        continue
                    # Coerce comma-formatted strings (e.g. "125,000") to numbers so
                    # HyperFormula arithmetic on the frontend doesn't get #VALUE!
                    val = cell.value
                    if isinstance(val, str):
                        stripped = val.replace(",", "")
                        try:
                            val = int(stripped) if "." not in stripped else float(stripped)
                        except (ValueError, TypeError):
                            pass
                    t_cell.value = val

        buf = io.BytesIO()
        template_wb.save(buf)
        return buf.getvalue()
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("Style merge failed, falling back to raw: %s", exc)
        return None


# ── 1. List assignments for this org/user ────────────────────────────────────


@router.get("/assignments", response_model=list[AssignmentForSubcontractor])
def list_my_assignments(
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return org-level template assignments for this user's org."""
    if not user.org_id:
        return []

    assignments = (
        db.query(TemplateAssignment)
        .filter(
            TemplateAssignment.submission_type == "template",
            TemplateAssignment.assigned_to_user_id.is_(None),
            TemplateAssignment.org_id == str(user.org_id),
        )
        .order_by(TemplateAssignment.assigned_at.desc())
        .all()
    )

    result = []
    for a in assignments:
        template_name = None
        if a.template_version_id:
            ver = db.query(TemplateVersion).filter(TemplateVersion.id == a.template_version_id).first()
            if ver:
                tmpl = db.query(Template).filter(Template.id == ver.template_id).first()
                template_name = tmpl.name if tmpl else None

        # Org-level: any submission by anyone in the org counts
        has_submission = (
            db.query(Submission)
            .filter(Submission.assignment_id == a.id)
            .first()
        ) is not None

        if a.status == "locked":
            org_status = "locked"
        else:
            org_status = "submitted" if has_submission else "pending"

        if a.assigned_to_user_id:
            # Legacy user-scoped assignment: org_id was the project_id
            pm = (
                db.query(ProjectMember)
                .filter(
                    ProjectMember.project_id == a.org_id,
                    ProjectMember.user_id == str(user.id),
                )
                .first()
            )
        else:
            # Org-based assignment: look up role by user_id across any project
            pm = (
                db.query(ProjectMember)
                .filter(ProjectMember.user_id == str(user.id))
                .order_by(ProjectMember.added_at.desc())
                .first()
            )

        result.append(
            AssignmentForSubcontractor(
                id=a.id,
                template_version_id=a.template_version_id,
                deadline=a.deadline,
                instructions=a.instructions,
                status=org_status,
                assigned_at=a.assigned_at,
                template_name=template_name,
                has_submission=has_submission,
                participant_role=pm.participant_role if pm else None,
            )
        )
    return result


# ── 2. Download the template Excel for an assignment ─────────────────────────


@router.get("/assignments/{assignment_id}/template-download")
def download_template(
    assignment_id: str,
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generate and serve an xlsx template from the PM's template schema."""
    if not user.org_id:
        raise HTTPException(status_code=400, detail="User has no org_id assigned")

    a = (
        db.query(TemplateAssignment)
        .filter(
            TemplateAssignment.id == assignment_id,
            _assignment_filter(user),
        )
        .first()
    )
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if not a.template_version_id:
        raise HTTPException(status_code=400, detail="No template linked to this assignment")

    ver = db.query(TemplateVersion).filter(TemplateVersion.id == a.template_version_id).first()
    if not ver:
        raise HTTPException(status_code=404, detail="Template version not found")

    tmpl = db.query(Template).filter(Template.id == ver.template_id).first()
    safe_name = (tmpl.name.replace(" ", "_") if tmpl else "template") + ".xlsx"

    # If an xlsx file was uploaded for this version, serve it directly
    if ver.file_path and Path(ver.file_path).exists():
        from fastapi.responses import Response
        with open(ver.file_path, "rb") as f:
            content = f.read()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_name}"',
                "Cache-Control": "no-store, no-cache, must-revalidate",
                "Pragma": "no-cache",
            },
        )

    # Fallback: generate a simple xlsx from the schema columns
    import openpyxl

    schema = ver.schema_json if isinstance(ver.schema_json, dict) else json.loads(ver.schema_json)
    columns = [col.get("name", col) for col in schema.get("columns", [])]
    if not columns:
        columns = list(schema.keys()) if isinstance(schema, dict) else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission"
    if columns:
        ws.append(columns)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FastAPIFileResponse(
        path=tmp.name,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── 3. Upload submission for an assignment ────────────────────────────────────


@router.post(
    "/assignments/{assignment_id}/submit",
    response_model=SubmissionResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(verify_csrf)],
)
async def submit_file(
    assignment_id: str,
    file: UploadFile = File(...),
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload an xlsx against an assignment. Resubmission replaces the prior file."""
    if not user.org_id:
        raise HTTPException(status_code=400, detail="User has no org_id assigned")

    a = (
        db.query(TemplateAssignment)
        .filter(
            TemplateAssignment.id == assignment_id,
            _assignment_filter(user),
        )
        .first()
    )
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if a.status == "locked":
        raise HTTPException(status_code=409, detail="Assignment is locked; no further submissions accepted")

    if not _is_admin(user):
        pm_q = db.query(ProjectMember).filter(
            ProjectMember.user_id == str(user.id),
            ProjectMember.participant_role == "focal",
        )
        if a.assigned_to_user_id:
            # Legacy user-scoped: org_id was the project_id
            pm_q = pm_q.filter(ProjectMember.project_id == a.org_id)
        pm = pm_q.first()
        if not pm:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only focal participants may submit files")

    settings = get_settings()
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are allowed")

    content = await file.read()
    max_bytes = settings.max_upload_size_mb * 1024 * 1024
    if len(content) > max_bytes:
        raise HTTPException(status_code=413, detail=f"File exceeds {settings.max_upload_size_mb} MB limit")
    _xlsx_magic = b"PK\x03\x04"
    _xls_magic = b"\xd0\xcf\x11\xe0"
    if not ((ext == ".xlsx" and content[:4] == _xlsx_magic) or (ext == ".xls" and content[:4] == _xls_magic)):
        raise HTTPException(status_code=400, detail="File does not appear to be a valid spreadsheet")

    sha = hash_file(content)
    org_dir = Path(settings.upload_dir) / "submissions" / str(user.org_id)
    org_dir.mkdir(parents=True, exist_ok=True)
    stored_path = org_dir / f"{sha}{ext}"

    # Merge submitted values into the original template so cell styles,
    # colors, merged cells and row heights are preserved for the admin view.
    styled_bytes = _merge_into_template(content, a.template_version_id, db)
    stored_path.write_bytes(styled_bytes if styled_bytes else content)

    existing = (
        db.query(Submission)
        .filter(
            Submission.assignment_id == assignment_id,
            Submission.submitted_by == str(user.id),
        )
        .first()
    )

    if existing:
        existing.file_path = str(stored_path)
        existing.file_name = file.filename or "submission.xlsx"
        existing.submitted_by = str(user.id)
        existing.status = "resubmitted"
        # Clear previous review so PIF sees it fresh
        existing.review_status = None
        existing.review_comment = None
        existing.reviewed_by = None
        existing.reviewed_at = None
        sub = existing
    else:
        sub = Submission(
            id=str(uuid.uuid4()),
            assignment_id=assignment_id,
            org_id=user.org_id,
            file_path=str(stored_path),
            file_name=file.filename or "submission.xlsx",
            status="submitted",
            submitted_by=str(user.id),
        )
        db.add(sub)

    a.status = "submitted"

    # ── Apply formulas if template version has any ────────────────────────
    if a.template_version_id:
        formulas = db.query(TemplateFormula).filter(TemplateFormula.template_version_id == a.template_version_id).all()
        if formulas:
            try:
                processed_bytes = FormulaExecutor.execute(content, formulas)
                processed_path = org_dir / f"{sha}_processed{ext}"
                processed_path.write_bytes(processed_bytes)
                if existing:
                    existing.processed_file_path = str(processed_path)
                else:
                    sub.processed_file_path = str(processed_path)
            except Exception as exc:
                import logging as _logging

                _logging.getLogger(__name__).warning(
                    "Formula execution failed for submission %s: %s", sub.id if not existing else existing.id, exc
                )

    db.commit()
    db.refresh(sub)
    return sub


# ── 4. List own submission history ───────────────────────────────────────────


@router.get("/submissions", response_model=list[SubmissionResponse])
def list_my_submissions(
    user: Profile = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return all submissions for assignments in the user's projects (org-level view)."""
    if not user.org_id:
        return []
    assignment_ids = [
        row[0] for row in db.query(TemplateAssignment.id)
        .filter(
            TemplateAssignment.org_id == str(user.org_id),
            TemplateAssignment.assigned_to_user_id.is_(None),
        ).all()
    ]
    if not assignment_ids:
        return []
    return (
        db.query(Submission)
        .filter(Submission.assignment_id.in_(assignment_ids))
        .order_by(Submission.submitted_at.desc())
        .all()
    )
